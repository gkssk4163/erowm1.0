# -*- coding: utf-8 -*-

import datetime
import json
import math
import os   # 파일삭제 시 경로관련..

from dateutil.relativedelta import relativedelta
from django.contrib import auth
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Sum, Count, Case, When, Q
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.dateformat import DateFormat

from accounting.common import getBudgetSumBySubsection, getBudgetSumByItem
from accounting.common import getLatestBudgetType, validBudgetDegree
from accounting.common import getTransactionListByName
from accounting.common import getTransactionSumBySubsection, getTransactionSumByItem
from accounting.view.subdivision import getSubdivisionList
from accounting.item import getItemList
from accounting.paragraph import getParagraphList
from accounting.subsection import getSubsectionList
from accounting.view.transaction import getTransactionList

from accounting.view.bankda import user_join_prs, user_withdraw, user_info_edit
from accounting.view.bankda import account_add, account_fix, account_del
from accounting.view.bankda import account_info_xml, account_list_partnerid_xml

from erowm import settings

from .common import session_info, subsection_info, item_info
from .forms import SignupForm, OwnerForm, BusinessForm, UserForm, SalesForm, AgencyForm, EditOwnerForm, AccountForm
from .forms import SubsectionForm, ParagraphForm, ItemForm, SubdivisionForm
from .forms import TblbankDirectForm, TransactionEditForm
from .models import Business_type
from .models import Owner, Business, Profile, Sales, Agency, Account
from .models import Subsection, Paragraph, Item, Subdivision
from .models import Transaction, TBLBANK, Budget, Deadline

from .domain.transaction import registTransaction, DeadlineCompletionError, NoTransactionHistoryForPreviousMonth

# Create your views here.

ACCOUNTANT = 1
OWNER = 2
SALES = 3
AGENCY = 4
LOCAL = 5

def main(request):
    user = request.user
    if not user.is_active:
        return redirect("login")
    else:
        request.session['master_login'] = False
        if user.profile.level_id >= LOCAL:
            return redirect("user_transform")
        elif user.profile.level_id >= SALES:
            return redirect("user_transform")
        else:
            return redirect("business_list")

def signup(request):
    if request.method == "POST":
        signupform = SignupForm(request.POST)
        ownerform = OwnerForm(request.POST)
        if signupform.is_valid() and ownerform.is_valid():
            # ** 뱅크다 계정등록 전에 erowm에 먼저 등록 필수 **
            # ** (뱅크다 먼저 등록 후 erowm 계정 생성 실패 시 뱅크다 계정 삭제 불가) **
            # erowm 계정등록
            user = signupform.save(commit=False)
            user.email = signupform.cleaned_data['email']
            owner = ownerform.save(commit=False)
            owner.email = signupform.cleaned_data['email']

            # 뱅크다 계정등록
            param = {
                'user_id': request.POST.get('username')
                , 'user_pw': request.POST.get('bankda_password')
                , 'user_name': request.POST.get('name')
                , 'user_tel': request.POST.get('cellphone')
                , 'user_email': request.POST.get('email')
            }
            result = user_join_prs(param)

            if result == "OK":  # 뱅크다 계정 정상등록
                # EROWM 계정등록
                try:
                    user.save()
                    Profile.objects.create(user=user, level_id=OWNER)
                    profile = Profile.objects.get(user=user)
                    owner.profile = profile
                    owner.save()
                except:
                    return render(request, "registration/bankda_join_error.html")
                return redirect("signup_done")
            else :
                return render(request, "registration/bankda_join_error.html")

    elif request.method == "GET":
        signupform = SignupForm()
        ownerform = OwnerForm()

    return render(request, "registration/signup.html", {
        "signupform": signupform,
        "ownerform": ownerform,
    })

def signup_done(request):
    return render(request, "registration/signup_done.html")

def login(request):
    # 해당 쿠키에 값이 없을 경우 None이 return 된다.
    # 쿠키에 username, password가 있으면 로그인 유지
    if request.COOKIES.get('username') is not None and request.COOKIES.get('password') is not None:
        username = request.COOKIES.get('username')
        password = request.COOKIES.get('password')
        user = auth.authenticate(request, username=username, password=password)
        if user is not None:
            auth.login(request, user)
            return redirect("main")
        else:
            return render(request, "registration/login.html")

    # POST
    elif request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]

        # 해당 user가 있으면 username, 없으면 None
        user = auth.authenticate(request, username=username, password=password)

        if user is not None:
            auth.login(request, user)
            response = redirect("main")

            if request.POST.get("keepLogin") == "on":   # 로그인유지 CHECK
                response.set_cookie('username', username)
                response.set_cookie('password', password)
            elif request.POST.get("saveInfo") == "on":    # 아이디저장 CHECK
                response.set_cookie('username', username)
            elif request.COOKIES.get('username') is not None:
                response.delete_cookie('username')
            return response
        else:
            return render(request, 'registration/login.html', {
                'username': username, 'error':'username or password is incorrect'})
    # GET
    else:
        username = request.COOKIES.get('username') if request.COOKIES.get('username') is not None else ""
        return render(request, 'registration/login.html', {'username': username})
    return render(request, 'registration/login.html')

def logout(request):
    response = redirect("main")
    response.delete_cookie('password')
    auth.logout(request)
    return response

@login_required(login_url='/')
def user_delete(request):
    if request.method == "POST":
        # 삭제권한 확인
        if request.user.profile.level_id < LOCAL:
            return HttpResponse("<script>alert('권한이 없습니다.');history.back();</script>")

        # 뱅크다 계정삭제
        param = {
            'user_id': request.POST.get('username')
            , 'user_pw': request.POST.get('bankda_password')
        }
        result = user_withdraw(param)

        if result == "OK":  # 뱅크다 계정 정상삭제
            # EROWM 계정삭제
            username = request.POST.get('username')
            user = get_object_or_404(User, username=username)
            user.delete()

        data = {'result': result}
        return JsonResponse(data, safe=False)  # safe=False 필수

@login_required(login_url='/')
def business_list(request):
    owner = request.user.profile.owner
    lists = Business.objects.filter(owner=owner)
    return render(request, 'accounting/business_list.html', {'lists': lists, 'master_login': request.session['master_login']})

@login_required(login_url='/')
def business_create(request):
    owner = request.user.profile.owner
    if request.method == "POST":
        form = BusinessForm(request.POST)
        if form.is_valid():
            business = form.save(commit=False)
            business.owner = owner
            if business.type3_id == '어린이집':
                business.session_month = '03'
            else :
                business.session_month = '01'
            business.save()
            return redirect('business_list')
    else:
        form = BusinessForm(initial={'place_name': owner.place_name, 'reg_number': owner.reg_number, 'owner_name': owner.name, 'owner_reg_number1': owner.owner_reg_number1, 'owner_reg_number2': owner.owner_reg_number2, 'type1': owner.type1, 'type2': owner.type2, 'cellphone': owner.cellphone, 'phone': owner.phone, 'fax': owner.fax, 'email': owner.email, 'zip_number': owner.zip_number, 'address': owner.address, 'detailed_address': owner.detailed_address})
    return render(request, 'accounting/business_edit.html', {'form': form})

@login_required(login_url='/')
def business_edit(request, pk):
    business = get_object_or_404(Business, pk=pk)
    owner = request.user.profile.owner
    if request.method == "POST":
        form = BusinessForm(request.POST, instance=business)

        # if form.is_valid() 안에 삭제 구현 시 기존파일 참조 안되는 경우 있어서 밖에 구현
        # 기존파일삭제 (파일을 변경하거나 삭제하는 경우 기존파일 삭제)
        if (business.ceo_stamp and request.POST.get('ceo_stamp_change') == 'true') \
                or request.POST.get('ceo_stamp-clear') == 'on':
            os.remove(os.path.join(settings.MEDIA_ROOT, business.ceo_stamp.name))
        if (business.manager_stamp and request.POST.get('manager_stamp_change') == 'true') \
                or request.POST.get('manager_stamp-clear') == 'on':
            os.remove(os.path.join(settings.MEDIA_ROOT, business.manager_stamp.name))
        if (business.business_stamp and request.POST.get('business_stamp_change') == 'true') \
                or request.POST.get('business_stamp-clear') == 'on':
            os.remove(os.path.join(settings.MEDIA_ROOT, business.business_stamp.name))

        if form.is_valid():
            business = form.save(commit=False)
            business.owner = owner
            if business.type3_id == '어린이집':
                business.session_month = '03'
            else :
                business.session_month = '01'

            if request.FILES.get('ceo_stamp') is not None:
                business.ceo_stamp = request.FILES.get('ceo_stamp')
            if request.FILES.get('manager_stamp') is not None:
                business.manager_stamp = request.FILES.get('manager_stamp')
            if request.FILES.get('business_stamp') is not None:
                business.business_stamp = request.FILES.get('business_stamp')

            business.save()
            return redirect('business_list')
    else:
        form = BusinessForm(instance=business)
    return render(request, 'accounting/business_edit.html', {'form': form, 'business': business})

@login_required(login_url='/')
def home(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    year = request.GET.get('year', today.strftime("%Y"))
    month = request.GET.get('month', today.strftime("%m"))
    inout_type = request.GET.get('type', 'input')

    sessionInfo = session_info(str(year), str(month), business.session_month)

    start_date = datetime.datetime.strptime(year+"-"+month+"-01", "%Y-%m-%d")
    end_date = start_date + relativedelta(months=1)

    #--마감자료
    try:
        deadline = Deadline.objects.get(business=business, year=year, month=month)
    except:
        deadline = None
    #--마감자료 END

    #--기타필요경비, 특별활동비, 장/단기 차입금 예산 결산 자료
    data_list2 = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3
    ).filter(Q(context__contains="기타필요경비")|Q(context__contains="특별활동비")|Q(context__contains="차입금")
    ).annotate(
        budget_amount=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(
                When(budget__year = sessionInfo['year'], then='budget__price'))))), 0)
    ).exclude(code=0).order_by(
        'paragraph__subsection__type',
        'paragraph__subsection__code',
        'paragraph__code',
        'code'
    )

    data_list = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3
    ).filter(Q(context__contains="기타필요경비")|Q(context__contains="특별활동비")|Q(context__contains="차입금")
    ).annotate(
        settlement_amount=Coalesce(Sum(Case(
            When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                    When(transaction__Bkdate__lt = end_date, then=Case(
                        When(
                            transaction__Bkoutput=0,
                            then='transaction__Bkinput'
                        ),
                        default='transaction__Bkoutput')))))))), 0)
    ).exclude(code=0).order_by(
        'paragraph__subsection__type',
        'paragraph__subsection__code',
        'paragraph__code',
        'code'
    )

    for idx, data in enumerate(data_list):
        data.budget_amount = math.ceil(data_list2[idx].budget_amount/12)
        try:
            data.rate_of_change = round(data.settlement_amount/data.budget_amount, 2)
        except:
            data.rate_of_change = '-'
    #--기타필요경비, 특별활동비, 장/단기 차입금 예산 결산 자료 END

    #--중복거래 내역
    duplicate_tr = Transaction.objects.filter(
        business=business, Bkdate__gte=start_date, Bkdate__lt=end_date
    ).values(
        'Bkdate', 'item__paragraph__subsection__type', 'item__context', 'item',
        'Bkjukyo', 'Bkinput', 'Bkoutput'
    ).annotate(price=Case(When(Bkinput__gt=0, then='Bkinput'), default='Bkoutput')
    ).annotate(count=Count('Bkdate')).filter(count__gte=2)
    #--중복거래 내역 END

    #--세입, 세출에 따른 목별 예산 결산 자료
    if inout_type == "input":
        filter_type = "수입"
        budget_type = "revenue"
    else:
        filter_type = "지출"
        budget_type = "expenditure"

    filter_budget_type = getLatestBudgetType(business, year, budget_type)

    budget_list = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type=filter_type
    ).annotate(
        budget_amount=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(
                When(budget__year = sessionInfo['year'], then=Case(When(
                    budget__type = filter_budget_type, then='budget__price'))))))), 0)
    ).exclude(code=0).order_by(
        'paragraph__subsection__type',
        'paragraph__subsection__code',
        'paragraph__code',
        'code'
    )

    item_list = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type=filter_type
    ).annotate(
        settlement_amount=Coalesce(Sum(Case(
            When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                    When(transaction__Bkdate__lt = end_date, then=Case(
                        When(
                            transaction__Bkoutput=0,
                            then='transaction__Bkinput'
                        ),
                        default='transaction__Bkoutput')))))))), 0)
    ).exclude(code=0).order_by(
        'paragraph__subsection__type',
        'paragraph__subsection__code',
        'paragraph__code',
        'code'
    )

    for idx, item in enumerate(item_list):
        item.budget_amount = math.ceil(budget_list[idx].budget_amount/12)
        try:
            item.rate_of_change = round(item.settlement_amount/item.budget_amount, 2)
        except:
            item.rate_of_change = '-'
    #--세입, 세출에 따른 목별 예산 결산 자료 END

    return render(request, 'accounting/home.html', {
        'business': business, 'y_range': range(today.year,1999,-1), 'm_range': range(1,13),
        'year': int(year), 'month': int(month), 'item_list': item_list, 'type': inout_type,
        'deadline': deadline, 'data_list': data_list, 'duplicate_tr': duplicate_tr })

@login_required(login_url='/')
def transform_business(request, pk):
    business = get_object_or_404(Business, pk=pk)
    request.session['business'] = business.pk
    return redirect('home')

@login_required(login_url='/')
def retransform_business(request):
    del request.session['business']
    return redirect('business_list')

@login_required(login_url='/')
def business_info(request):
    owner = request.user.profile.owner
    business = get_object_or_404(Business, pk=request.session['business'])
    if business.owner != owner:
        return redirect('business_list')
    return render(request, 'accounting/business_info.html', {
        'business': business, 'business_info_page': 'active', 'business_management': 'active', 'master_login': request.session['master_login']
    })

@login_required(login_url='/')
def account_list(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    lists = Account.objects.filter(business=business)

    for list in lists:
        data = {'bkacctno': list.account_number}
        list.bankda = account_info_xml(data)

    return render(request, 'accounting/account_list.html', {
        'lists': lists, 'business': business, 'accounting_management': 'active',
        'account_list': 'active', 'master_login': request.session['master_login'] })

@login_required(login_url='/')
def account_create(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    main_acct = Account.objects.filter(business=business, main=True).count()
    if request.method == "POST":
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save(commit=False)
            if main_acct:   # 주계좌가 있는 경우
                account.main = False    # 신규 등록계좌 주계좌로 설정하지 않음

            # 뱅크다 계좌등록
            Bjumin = business.reg_number.split("-")
            param = {
                'user_id': request.user.username
                , 'user_pw': request.user.profile.owner.bankda_password
                , 'bkdiv': request.POST.get('bkdiv')
                , 'bkcode': request.POST.get('bank')
                , 'bkacctno': request.POST.get('account_number')
                , 'bkacctpno_pw': request.POST.get('account_pw')
                , 'Mjumin_1': business.owner_reg_number1
                , 'Bjumin_1': Bjumin[0]
                , 'Bjumin_2': Bjumin[1]
                , 'Bjumin_3': Bjumin[2]
                , 'webid': request.POST.get('webid')
                , 'webpw': request.POST.get('webpw')
                , 'renames': request.POST.get('renames')
            }
            result = account_add(param)

            if result == "OK":  # 뱅크다 계좌 정상등록
                # EROWM 계좌등록
                account.save()
                return redirect('account_list')
            else:   # 뱅크다 계좌등록 오류
                return render(request, "accounting/bankda_error.html")
    else:
        form = AccountForm(initial={'business': business})
    return render(request, 'accounting/account_edit.html', {
        'form': form, 'business': business, "editType": "create"})

@login_required(login_url='/')
def account_edit(request, pk):
    account = get_object_or_404(Account, pk=pk)
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "POST":
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            account = form.save(commit=False)

            # 뱅크다 계좌수정(은행명/계좌번호 수정불가)
            Bjumin = business.reg_number.split("-")
            param = {
                'user_id': request.user.username
                , 'user_pw': request.user.profile.owner.bankda_password
                , 'bkdiv': request.POST.get('bkdiv')
                , 'bkcode': request.POST.get('bank')
                , 'bkacctno': request.POST.get('account_number')
                , 'bkacctpno_pw': request.POST.get('account_pw')
                , 'Mjumin_1': business.owner_reg_number1
                , 'Bjumin_1': Bjumin[0]
                , 'Bjumin_2': Bjumin[1]
                , 'Bjumin_3': Bjumin[2]
                , 'webid': request.POST.get('webid')
                , 'webpw': request.POST.get('webpw')
                , 'renames': request.POST.get('renames')
            }
            result = account_fix(param)

            if result == "OK":  # 뱅크다 계좌 정상수정
                # EROWM 계좌수정
                account.save()
                return redirect('account_list')
            else:   # 뱅크다 계좌수정 오류
                return render(request, "accounting/bankda_error.html")
    else:
        form = AccountForm(instance=account)
    return render(request, 'accounting/account_edit.html', {'acctpk':pk, 'form': form, 'business': business})

@login_required(login_url='/')
def account_delete(request):
    user = get_object_or_404(User, pk = request.POST.get('user_pk'))
    account = get_object_or_404(Account, pk = request.POST.get('account_pk'))

    # 삭제권한 확인
    if request.user.profile.level_id < LOCAL:
        # 사용자페이지 삭제구현 시 계좌소유자만 삭제할 수 있도록 함(기능 확인필요)
        # if account.business.owner.profile.user != request.user:
        return HttpResponse("<script>alert('권한이 없습니다.');history.back();</script>")

    param = {
        'user_id': user.username
        , 'user_pw': user.profile.owner.bankda_password
        , 'bkacctno': account.account_number
    }
    result = account_del(param)

    if result == "OK":  # 뱅크다 계좌 정상삭제
        # EROWM 계좌삭제
        account.delete()

    data = {'result': result}
    return JsonResponse(data, safe=False)  # safe=False 필수

@login_required(login_url='/')
def account_disconnect(request):
    user = get_object_or_404(User, pk = request.POST.get('user_pk'))
    account = get_object_or_404(Account, pk = request.POST.get('account_pk'))

    # 연결해제권한 확인
    if request.user.profile.level_id < LOCAL:
        return HttpResponse("<script>alert('권한이 없습니다.');history.back();</script>")

    param = {
        'user_id': user.username
        , 'user_pw': user.profile.owner.bankda_password
        , 'bkacctno': account.account_number
    }
    result = account_del(param)

    data = {'result': result}
    return JsonResponse(data, safe=False)  # safe=False 필수

@login_required(login_url='/')
def user_list(request):
    if request.user.is_staff:
        owners = Owner.objects.all()
        where = request.GET.get('where', '')
        keyword = request.GET.get('keyword', '')
        sales = Sales.objects.all()
        if where == 'place_name'  and keyword:
            owners = owners.filter(place_name__icontains=keyword)
        elif where == 'name' and keyword:
            owners = owners.filter(name__icontains=keyword)
        elif where == 'sales' and keyword:
            owners = owners.filter(sales__name__icontains=keyword)
    else:
        return redirect('business_list')
    return render(request, 'accounting/user_list.html', {
        'owners': owners, 'sales': sales, 'user_list_page': 'active'
    })

@login_required(login_url='/')
def user_transform(request):
    user = request.user
    if user.is_staff:
        owners = Owner.objects.all()
    elif user.profile.level_id == SALES:
        owners = Owner.objects.filter(sales=user.profile.sales)
    elif user.profile.level_id == AGENCY:
        sales_list = list()
        sales_set = Sales.objects.filter(agency=user.profile.agency)
        for sales in sales_set:
            sales_list.append(sales.id)
        owners = Owner.objects.filter(sales_id__in=sales_list)
    else:
        return redirect('business_list')

    where = request.GET.get('where', '')
    keyword = request.GET.get('keyword', '')
    if where == 'place_name'  and keyword:
        owners = owners.filter(place_name__icontains=keyword)
    elif where == 'name' and keyword:
        owners = owners.filter(name__icontains=keyword)

    owner_exists = owners.exists()

    return render(request, 'accounting/user_transform.html', {
        'owners': owners, 'owner_exists': owner_exists, 'user_transform_page': 'active'})

@login_required(login_url='/')
def transform(request, pk):
    user = get_object_or_404(User, pk=pk)
    if (request.user.profile.level_id >= SALES) and user.is_active:
        master_username = request.user.username
        auth.login(request, user)
        if user is not None:
            request.session['master_login'] = True
            request.session['username'] = master_username
            return redirect('business_list')
    return redirect('user_transform')

@login_required(login_url='/')
def retransform(request):
    if request.session['master_login'] == True:
        username = request.session['username']
        user = User.objects.get(username=username)
        auth.login(request, user)
        if user is not None:
            request.session['username'] = user.username
            request.session['master_login'] = False
            return redirect('user_transform')
    return redirect('business_list')

@login_required(login_url='/')
def sales_list(request):
    user = request.user
    if user.is_staff:
        sales = Sales.objects.all()
        agency = Agency.objects.all()
        where = request.GET.get('where', '')
        keyword = request.GET.get('keyword', '')
        if where == 'name'  and keyword:
            sales = sales.filter(name__icontains=keyword)
        elif where == 'agency' and keyword:
            sales = sales.filter(agency__name__contains=keyword)
    else:
        return redirect('business_list')
    return render(request, 'accounting/sales_list.html', {
        'sales': sales, 'agency': agency, 'sales_list_page': 'active'
    })

@login_required(login_url='/')
def sales_create(request):
    if request.method == "POST":
        userform = UserForm(request.POST)
        form = SalesForm(request.POST)
        if userform.is_valid() and form.is_valid():
            user = userform.save(commit=False)
            user.email = userform.cleaned_data['email']
            user.save()
            Profile.objects.create(user=user, level_id=SALES)
            profile = Profile.objects.get(user=user)
            sales = form.save(commit=False)
            sales.profile = profile
            sales.email = userform.cleaned_data['email']
            sales.save()
            return redirect('sales_list')
    else:
        userform = UserForm()
        form = SalesForm()
    return render(request, 'accounting/sales_edit.html', {'form': form, 'userform': userform, 'sales_list_page': 'active'})

@login_required(login_url='/')
def sales_edit(request, pk):
    sales = get_object_or_404(Sales, pk=pk)
    if request.method == "POST":
        form = SalesForm(request.POST, instance=sales)
        if form.is_valid():
            form.save()
            return redirect('sales_list')
    else:
        form = SalesForm(instance=sales)
    return render(request, 'accounting/sales_edit.html', {'form': form, 'sales_list_page': 'active'})

@login_required(login_url='/')
def sales_delete(request, pk):
    sales = get_object_or_404(Sales, pk=pk)
    sales.delete()
    return redirect('sales_list')

@login_required(login_url='/')
def sales_change(request):
    pk_list = request.POST.getlist('check_list[]')
    if request.method == "POST":
        auth_selected = request.POST.get('select_auth', '')
        sales_selected = request.POST.get('select_sales', '')
        for pk in pk_list:
            owner = get_object_or_404(Owner, pk=pk)
            owner.is_demo = auth_selected
            owner.sales_id = sales_selected
            owner.save()
    return redirect('user_list')

@login_required(login_url='/')
def agency_list(request):
    user = request.user
    if user.is_staff:
        agency = Agency.objects.all()
        where = request.GET.get('where', '')
        keyword = request.GET.get('keyword', '')
        if where == 'name'  and keyword:
            agency = agency.filter(name__icontains=keyword)
        elif where == 'owner_name' and keyword:
            agency = agency.filter(owner_name__icontains=keyword)
    else:
        return redirect('business_list')
    return render(request, 'accounting/agency_list.html', {
        'agency': agency, 'agency_list_page': 'active'
    })

@login_required(login_url='/')
def agency_create(request):
    if request.method == "POST":
        userform = UserForm(request.POST)
        form = AgencyForm(request.POST)
        if userform.is_valid() and form.is_valid():
            user = userform.save(commit=False)
            user.email = userform.cleaned_data['email']
            user.save()
            Profile.objects.create(user=user, level_id=AGENCY)
            profile = Profile.objects.get(user=user)
            agency = form.save(commit=False)
            agency.profile = profile
            agency.email = userform.cleaned_data['email']
            agency.save()
            return redirect('agency_list')
    else:
        userform = UserForm()
        form = AgencyForm()
    return render(request, 'accounting/agency_edit.html', {'form': form, 'userform': userform, 'agency_list_page': 'active'})

@login_required(login_url='/')
def agency_edit(request, pk):
    agency = get_object_or_404(Agency, pk=pk)
    if request.method == "POST":
        form = AgencyForm(request.POST, instance=agency)
        if form.is_valid():
            form.save()
            return redirect('agency_list')
    else:
        form = AgencyForm(instance=agency)
    return render(request, 'accounting/agency_edit.html', {'form': form, 'agency_list_page': 'active'})

@login_required(login_url='/')
def agency_delete(request, pk):
    agency = get_object_or_404(Agency, pk=pk)
    agency.delete()
    return redirect('agency_list')

@login_required(login_url='/')
def agency_change(request):
    pk_list = request.POST.getlist('check_list[]')
    if request.method == "POST":
        selected_agency = request.POST.get('select_agency', '')
        for pk in pk_list:
            sales = get_object_or_404(Sales, pk=pk)
            sales.agency_id = selected_agency
            sales.save()
    return redirect('sales_list')

@login_required(login_url='/')
def mypage(request):
    owner = request.user.profile.owner
    return render(request, 'accounting/mypage.html', {
        'owner': owner, 'mypage': 'active',
    })

@login_required(login_url='/')
def mypage_edit(request):
    user = request.user
    owner = request.user.profile.owner
    if request.method == "POST":
        form = EditOwnerForm(request.POST, instance=owner)
        if form.is_valid():
            # EROWM 계정 정보세팅
            owner = form.save(commit=False)

            # 뱅크다 계정수정
            param = {
                'user_id': request.POST.get('username')
                , 'user_pw': request.POST.get('bankda_password')
                , 'user_name': request.POST.get('name')
                , 'user_tel': request.POST.get('cellphone')
                , 'user_email': request.POST.get('email')
                , 'user_pw_new': request.POST.get('user_pw_new')
            }
            result = user_info_edit(param)
            if result == "OK":  # 뱅크다 계정 정상수정
                # EROWM 계정등록
                if request.POST.get('user_pw_new'): # 뱅크다 PW는 PW가 입력된 경우만 변경
                    owner.bankda_password = request.POST.get('user_pw_new')
                owner.save()
                return redirect('mypage')
            else:  # 뱅크다 계정수정 오류
                return render(request, "accounting/bankda_error.html")
    else:
        form = EditOwnerForm(instance=owner)
    return render(request, 'accounting/mypage_edit.html', {'user': user, 'form': form})

@login_required(login_url='/')
def bankda_join(request):
    owners = Owner.objects.all()
    return render(request, 'accounting/bankda_join.html', {
        'owners': owners, 'bankda_page': 'active',
    })

@login_required(login_url='/')
def bankda_account(request):
    accounts = account_list_partnerid_xml()
    return render(request, 'accounting/bankda_account.html', {
        'accounts': accounts, 'bankda_page': 'active', 'bankda_account_page': 'active',
    })

@login_required(login_url='/')
def transaction_history(request):
    today = datetime.datetime.now()
    business = get_object_or_404(Business, pk=request.session['business'])
    acct_list = Account.objects.filter(business=business)
    try:
        main_acctid = Account.objects.get(business=business, main=True).id
    except:
        main_acctid = 0

    acctid = request.GET.get('acctid', main_acctid)
    year = request.GET.get('year', DateFormat(today).format("Y"))
    month = request.GET.get('month', DateFormat(today).format("m"))
    page = request.GET.get('page', 1)
    page2 = request.GET.get('page2', 1)

    sessionInfo = session_info(year, month, business.session_month)

    if main_acctid == 0:
        account_number = None
    else:
        account_number = get_object_or_404(Account, business=business, id=acctid).account_number

    start_date = datetime.datetime.strptime(year+'-'+month+'-01', '%Y-%m-%d')
    end_date = start_date + relativedelta(months=1)

    input_items = item_info(sessionInfo['year'], business.type3, 'i')
    output_items = item_info(sessionInfo['year'], business.type3, 'o')
    input_subsections = subsection_info(sessionInfo['year'], business.type3, 'i')

    selected_check_list = []
    for idx, val in enumerate(request.POST.getlist('tr_check_list')):
        selected_check_list.append(int(val))
    item_list = request.POST.getlist('item_list')
    selected_item_list = []
    subdivision_list = []
    for idx, val in enumerate(item_list):
        if val:
            selected_item_list.append(int(val))
        else:
            selected_item_list.append(0)
        subdivisions = Subdivision.objects.filter(business = business, item__id = selected_item_list[idx])
        subdivision_list.append(subdivisions)

    input_subsection_list = request.POST.getlist('input_subsection_list')
    selected_subsection_list = []
    relative_item_list = []
    for idx, val in enumerate(input_subsection_list):
        if val:
            selected_subsection_list.append(int(val))
        else:
            selected_subsection_list.append(0)
        relative_item = Item.objects.filter(paragraph__subsection__id = selected_subsection_list[idx])
        relative_item_list.append(relative_item)

    subdivision_lists = request.POST.getlist('subdivision_list')
    selected_subdivision_list = []
    for idx, val in enumerate(subdivision_lists):
        if val:
            selected_subdivision_list.append(int(val))
        else:
            selected_subdivision_list.append(0)

    input_subdivision_list = request.POST.getlist('input_subdivision_list')
    selected_input_subdivision_list = []
    for idx, val in enumerate(input_subdivision_list):
        if val:
            selected_input_subdivision_list.append(int(val))
        else:
            selected_input_subdivision_list.append(0)

    data_list = TBLBANK.objects.filter(business=business, Bkacctno=account_number, Bkdate__gte=start_date, Bkdate__lt=end_date).exclude(item__code=0).order_by('-Bkdate', '-Bkid', 'Bkdivision')
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date).exclude(item__code=0).order_by('-Bkdate', '-id', 'Bkdivision')
    #print(start_date, data_list[0].Bkdate)

    #--수입/지출계
    try:
        premonth_transfer_price = transaction_list.get(Bkdate=start_date, Bkdivision=0).Bkjango
    except:
        premonth_transfer_price = 0

    total_input = 0
    total_output = 0

    for transaction in transaction_list:
        if transaction.Bkinput != 0:
            total_input += transaction.Bkinput
        elif transaction.Bkoutput != 0:
            total_output += transaction.Bkoutput
    total_input -= premonth_transfer_price

    try:
        jango = transaction_list.first().Bkjango
    except:
        jango = 0

    paginator = Paginator(data_list, 10)
    paginator2 = Paginator(transaction_list, 10)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = Paginator.page(paginator.num_pages)

    try:
        data2 = paginator2.page(page2)
    except PageNotAnInteger:
        data2 = paginator2.page(1)
    except EmptyPage:
        data2 = Paginator.page(paginator2.num_pages)
    return render(request, 'accounting/transaction_history.html', {
        'selected_check_list': selected_check_list, 'total_input': total_input, 'total_output': total_output, 'jango': jango, 'premonth_transfer_price': premonth_transfer_price, 'year': int(year), 'month': int(month), 'year_range': range(int(DateFormat(today).format("Y")), 1999, -1), 'month_range': range(1,13), 'acctid': int(acctid), 'page': page, 'page2': page2, 'acct_list': acct_list, 'selected_subdivision_list': selected_subdivision_list, 'selected_input_subdivision_list': selected_input_subdivision_list, 'relative_item_list': relative_item_list, 'input_subsections': input_subsections, 'selected_subsection_list': selected_subsection_list, 'selected_item_list': selected_item_list, 'data': data, 'data2': data2, 'input_items': input_items, 'output_items': output_items, 'subdivision_list': subdivision_list, 'business': business, 'accounting_management': 'active', 'transaction_history_page': 'active', 'master_login': request.session['master_login'],
    })

@login_required(login_url='/')
def transaction_list(request):
    business = get_object_or_404(Business, pk=request.session['business'])

    if request.method == "GET":
        today = datetime.datetime.now()
        firstDay = today.replace(day=1)
        nextMonth = firstDay + relativedelta(months=1)
        lastDay = nextMonth - relativedelta(days=1)

        start_date = request.GET.get('start_date', firstDay)
        end_date = request.GET.get('end_date', lastDay)
        keyword = request.GET.get('keyword', "")

        param = {
            'start_date': DateFormat(start_date).format("Y-m-d")
            ,'end_date': DateFormat(end_date).format("Y-m-d")
            ,'keyword': keyword
        }

        return render(request, 'accounting/transaction/list.html', {
            'business': business, 'master_login': request.session['master_login'],
            'accounting_management': 'active', 'transaction_list_page': 'active',
            'param': param
        })
    else:
        param = {
            'business': business
            ,'start_date': request.POST.get('start_date')
            ,'end_date': request.POST.get('end_date')
            ,'keyword': request.POST.get('keyword')
            ,'type': request.POST.get('type')
            ,'subsection': request.POST.get('subsection')
            ,'paragraph': request.POST.get('paragraph')
            ,'item': request.POST.get('item')
            ,'subdivision': request.POST.get('subdivision')
            ,'codeYN': 'N'
        }

        print(param)

        transaction = getTransactionList(param)  # QuerySet
        data = list(transaction.values())  # JsonResponse를 사용하여 전달하기 위해 QuerySet을 list 타입으로 변경

        for d in data:
            # item = model_to_dict(Item.objects.get(pk = d['item_id']))
            item = Item.objects.get(pk = d['item_id'])
            d['io_type'] = item.paragraph.subsection.type
            d['spi_code'] = str(item.paragraph.subsection.code) + str(item.paragraph.code) + str(item.code)
            d['item_context'] = item.context
            print(d['subdivision_id'])
            if d['subdivision_id'] is not None:
                subdivision = Subdivision.objects.get(pk=d['subdivision_id'])
                d['subdivision_context'] = subdivision.context
            else:
                d['subdivision_context'] = ""

        return JsonResponse(data, safe=False)  # safe=False 필수

# @login_required(login_url='/')
# def regist_transaction(request):
#     business = get_object_or_404(Business, pk=request.session['business'])
#     acctid = request.POST.get('acctid')
#     acct = get_object_or_404(Account, business=business, id=acctid)
#     Mid = request.user.username
#     year = request.POST.get('year')
#     month = request.POST.get('month')
#     page = request.POST.get('page')
#     page2 = request.POST.get('page2')
#     tr_check_list = request.POST.getlist('tr_check_list')
#     tr_list = request.POST.getlist('transaction_list')
#     Bkdate_list = request.POST.getlist('Bkdate_list')
#     Bkjukyo_list = request.POST.getlist('Bkjukyo_list')
#     item_list = request.POST.getlist('item_list')
#     subdivision_list = request.POST.getlist('subdivision_list')
#     relative_subsection_list = request.POST.getlist('input_subsection_list')
#     relative_item_list = request.POST.getlist('input_subdivision_list')
#
#     # for check in tr_check_list:
#     #     if int(check) != 0 and item_list[int(check)] == '':
#     #         return HttpResponse("<script>alert('계정명을 선택해주세요.');history.back();</script>")
#
#     tr_check_list.reverse()
#
#     if request.method == "POST":
#         print("views.py regist_transaction")
#         # from django.db import transaction
#         # with transaction.atomic():
#         try:
#             for check in tr_check_list:
#                 Bkid = tr_list[int(check)]
#                 tblbank_tr = TBLBANK.objects.get(Bkid=Bkid)
#                 try:
#                     subdivision = Subdivision.objects.get(id=subdivision_list[int(check)])
#                 except Subdivision.DoesNotExist:
#                     subdivision = None
#                 try:
#                     relative_subsection = Subsection.objects.get(id=relative_subsection_list[int(check)])
#                 except Subsection.DoesNotExist:
#                     relative_subsection = None
#                 try:
#                     relative_item = Item.objects.get(id=relative_item_list[int(check)])
#                 except Item.DoesNotExist:
#                     relative_item = None
#
#                 tr = Transaction(
#                     Bkid=Bkid,
#                     Mid=Mid,
#                     business=business,
#                     Bkacctno=acct.account_number,
#                     Bkname=acct.bank.name,
#                     Bkdate=datetime.datetime.strptime(Bkdate_list[int(check)], "%Y-%m-%d"),
#                     Bkjukyo=Bkjukyo_list[int(check)],
#                     Bkinput=tblbank_tr.Bkinput,
#                     Bkoutput=tblbank_tr.Bkoutput,
#                     item=Item.objects.get(id=item_list[int(check)]),
#                     subdivision=subdivision,
#                     relative_subsection=relative_subsection,
#                     relative_item=relative_item
#                 )
#                 registTransaction(business, tr)
#         except DeadlineCompletionError as e:
#             return HttpResponse("<script>alert('" + e.__str__() + "');history.back();</script>")
#         except NoTransactionHistoryForPreviousMonth as e:
#             return HttpResponse("<script>alert('" + e.__str__() + "');history.back();</script>")
#
#     response = redirect('transaction_history')
#     response['Location'] += '?page='+page+'&page2='+page2+'&year='+year+'&month='+month+'&acctid='+acctid
#     return response

@login_required(login_url='/')
def transaction_delete(request):
    Mid = request.user.username
    business = get_object_or_404(Business, pk=request.session['business'])
    check_list = request.POST.getlist('check_list[]')
    acctid = request.POST.get('acctid')
    year = request.POST.get('year')
    month = request.POST.get('month')
    page = request.POST.get('page')
    page2 = request.POST.get('page2')

    try:
        close = Deadline.objects.get(business=business, year=year, month=month)
        if close.regdatetime:
            return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
    except:
        pass

    for Bkid in check_list:
        transaction_list = Transaction.objects.filter(Bkid=Bkid).exclude(Bkdivision=0)
        tblbank_list = TBLBANK.objects.filter(Bkid=Bkid).exclude(Bkdivision=0)
        for tr in transaction_list:
            #잔고계산
            a_month_later = DateFormat(tr.Bkdate + relativedelta(months=1)).format("Y-m-01")
            update_list = Transaction.objects.filter(business=business, Bkdate__gte=tr.Bkdate, Bkdate__lt=a_month_later)
            for update in update_list:
                if (update.Bkdate == tr.Bkdate and update.id > tr.id) or update.Bkdate != tr.Bkdate:
                    if tr.Bkinput != 0:
                        update.Bkjango = update.Bkjango - tr.Bkinput
                    elif tr.Bkoutput != 0:
                        update.Bkjango = update.Bkjango + tr.Bkoutput
                    update.save()
            tr.delete()
        for tblbank in tblbank_list:
            if tblbank.Bkdivision == 1 and tblbank.direct == False:
                tblbank.regdatetime = None
                tblbank.item = None
                tblbank.relative_item = None
                tblbank.relative_subsection = None
                tblbank.subdivision = None
                tblbank.sub_Bkjukyo = None
                tblbank.save()
            else:
                tblbank.delete()

    response = redirect('transaction_history')
    response['Location'] += '?page='+page+'&page2='+page2+'&year='+year+'&month='+month+'&acctid='+acctid
    return response

@login_required(login_url='/')
def spi_list(request):
    institution_name = request.GET.get('institution_name', '어린이집')
    institution_list = Business_type.objects.all()

    subsection_list = Subsection.objects.filter(institution=institution_name).annotate(count=Count('paragraph__item')).exclude(count=0).order_by('year','institution','type','code')
    #paragraph_list = Paragraph.objects.all().annotate(count=Count('item')).exclude(count=0).order_by('code')
    #item_list = Item.objects.all().order_by('code')
    paragraph_list = Paragraph.objects.filter(subsection__institution=institution_name).annotate(count=Count('item')).exclude(count=0).order_by('code')
    item_list = Item.objects.filter(paragraph__subsection__institution=institution_name).order_by('code')
    return render(request, 'accounting/spi_list.html', {
        'setting_page': "active", 'spi_list_page': "active",
        'subsection_list': subsection_list, 'paragraph_list': paragraph_list,
        'item_list': item_list, 'institution_list': institution_list})

@login_required(login_url='/')
def subsection_create(request):
    if request.method == "POST":
        form = SubsectionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = SubsectionForm()
    return render(request, 'accounting/subsection_edit.html', {'form': form})

@login_required(login_url='/')
def subsection_edit(request, pk):
    subsection = Subsection.objects.get(pk=pk)
    if request.method == "POST":
        form = SubsectionForm(request.POST, instance=subsection)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = SubsectionForm(instance=subsection)
    return render(request, 'accounting/subsection_edit.html', {'form': form})

@login_required(login_url='/')
def paragraph_create(request):
    if request.method == "POST":
        form = ParagraphForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = ParagraphForm()
    return render(request, 'accounting/paragraph_edit.html', {'form': form})

@login_required(login_url='/')
def paragraph_edit(request, pk):
    paragraph = Paragraph.objects.get(pk=pk)
    if request.method == "POST":
        form = ParagraphForm(request.POST, instance=paragraph)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = ParagraphForm(instance=paragraph)
    return render(request, 'accounting/paragraph_edit.html', {'form': form})

@login_required(login_url='/')
def item_create(request):
    if request.method == "POST":
        form = ItemForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = ItemForm()
    return render(request, 'accounting/item_edit.html', {'form': form})

@login_required(login_url='/')
def item_edit(request, pk):
    item = Item.objects.get(pk=pk)
    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('spi_list')
    else:
        form = ItemForm(instance=item)
    return render(request, 'accounting/item_edit.html', {'form': form})

@login_required(login_url='/')
def subdivision_list(request):
    if request.method == "POST":
        item = request.POST.get('item')

        subdivision = getSubdivisionList(request.session['business'], item)  # QuerySet
        data = list(subdivision.values())  # JsonResponse를 사용하여 전달하기 위해 QuerySet을 list 타입으로 변경

        return JsonResponse(data, safe=False)  # safe=False 필수
    else:
        business = get_object_or_404(Business, pk=request.session['business'])
        today = datetime.datetime.now()
        this_year = DateFormat(today).format("Y")
        this_month = DateFormat(today).format("m")
        sessionInfo = session_info(this_year, this_month, business.session_month)
        year = int(request.GET.get('year', sessionInfo['year']))
        subdivisions = Subdivision.objects.filter(
            business = business
            ,item__paragraph__subsection__year = year
        ).order_by(
                'item__paragraph__subsection__type', 'item__paragraph__subsection__code'
                , 'item__paragraph__code', 'item__code', 'code')
        return render(request, 'accounting/subdivision_list.html', {
            'business': business, 'subdivisions': subdivisions
            ,'master_login': request.session['master_login']
            ,'accounting_management': 'active', 'subdivision_list': 'active'
            ,'year': year, 'year_range': range(int(sessionInfo['year']), 2018, -1)
            })

@login_required(login_url='/')
def subdivision_create(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = DateFormat(today).format("Y")
    this_month = DateFormat(today).format("m")
    sessionInfo = session_info(this_year, this_month, business.session_month)
    year = int(request.GET.get('year', sessionInfo['year']))
    if request.method == "POST":
        form = SubdivisionForm(request.POST)
        if form.is_valid():
            subdivision = form.save(commit=False)
            subdivision.business = business
            subdivision.save()
            return redirect('subdivision_list')
    else:
        form = SubdivisionForm(initial={'business': business})
        form.fields['item'].queryset = Item.objects.filter(
            paragraph__subsection__institution = business.type3
            ,paragraph__subsection__year = year
        ).exclude(code=0)
    return render(request, 'accounting/subdivision_edit.html', {
        'business': business, 'form': form
        ,'accounting_management': 'active', 'subdivision_list': 'active'
        ,'year': year, 'year_range': range(int(sessionInfo['year']), 2018, -1)
        })

@login_required(login_url='/')
def select_item(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = DateFormat(today).format("Y")
    this_month = DateFormat(today).format("m")
    sessionInfo = session_info(this_year, this_month, business.session_month)
    year = request.POST.get('year', sessionInfo['year'])
    item_list = Item.objects.filter(
        paragraph__subsection__institution = business.type3
        ,paragraph__subsection__year = year
    ).exclude(code=0)

    option_html='<option value="" selected="">---------</option>\n'
    for item in item_list:
        option_html += '<option value="'+str(item.pk)+'">'+str(item)+'</option>\n'
    print(option_html)

    return HttpResponse(json.dumps({'option_html': option_html}), content_type="application/json")

@login_required(login_url='/')
def other_settings(request):
    return render(request, 'accounting/other_settings.html')

@login_required(login_url='/')
def database_syn(request):
    from .models import Setting
    tr_num = Setting.objects.get(name="transaction_num")

    from django.db import connections
    query = """SELECT * FROM TBLBANK WHERE Bkid > """+str(tr_num.value)

    with connections['default'].cursor() as cursor:
        cursor.execute(query)
        columns = [ col[0] for col in cursor.description ]
        data_list = [ dict(zip(columns,row)) for row in cursor.fetchall() ]

    try:
        num = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    except:
        num = 1

    for data in data_list:
        business = Business.objects.get(account__account_number=data['Bkacctno'])
        TBLBANK.objects.create(
            Bkid=num,
            Bkdivision=1,
            Mid=data['Mid'],
            Bkacctno=data['Bkacctno'],
            Bkname=data['Bkname'],
            Bkdate=data['Bkdate'][:4]+'-'+data['Bkdate'][4:6]+'-'+data['Bkdate'][6:],
            Bkjukyo=data['Bkjukyo'],
            Bkinput=data['Bkinput'],
            Bkoutput=data['Bkoutput'],
            Bkjango=data['Bkjango'],
            business=business
        )
        tr_num.value = data['Bkid']
        tr_num.save()
        num += 1

    return redirect('other_settings')

@login_required(login_url='/')
def popup_change_main_account(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    acct_list = Account.objects.filter(business=business)

    return render(request, 'accounting/popup_change_mainaccount.html', {'acct_list' :acct_list})

@login_required(login_url='/')
def change_main_account(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    acct_list = Account.objects.filter(business=business)
    if request.method == "POST":
        main_acctid = int(request.POST.get('acctid'))
        for acct in acct_list:
            if acct.id == main_acctid:
                acct.main = True
            else:
                acct.main = False
            acct.save()
    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def popup_transaction_division(request, Bkid):
    business = get_object_or_404(Business, pk=request.session['business'])
    month = request.GET.get('month')
    tblbank_tr = TBLBANK.objects.get(Bkid=Bkid)
    total = 0
    if tblbank_tr.Bkinput:
        total = tblbank_tr.Bkinput
        filter_type = "수입"
    elif tblbank_tr.Bkoutput:
        total = tblbank_tr.Bkoutput
        filter_type = "지출"

    sessionInfo = session_info(str(tblbank_tr.Bkdate.year), str(tblbank_tr.Bkdate.month), business.session_month)

    item_list=Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution=business.type3,
        paragraph__subsection__type = filter_type).exclude(code=0)
    return render(request, 'accounting/popup_transaction_division.html', {'transaction': tblbank_tr, 'Bkid': Bkid, 'total': total, 'month': month, 'item_list': item_list})

@login_required(login_url='/')
def select_subdivision(request):
    from django.core import serializers
    business = get_object_or_404(Business, pk=request.session['business'])
    item_id = request.POST.get('item_id')
    subdivision_list = Subdivision.objects.filter(item__id=item_id, business=business)
    data_json = serializers.serialize('json', subdivision_list)
    print(data_json)
    return HttpResponse(data_json, content_type="application/json")

@login_required(login_url='/')
def add_row(request):
    context = {}
    page = request.POST.get('page', None)
    if page == "popup_transaction_division":
        Bkid = request.POST.get('Bkid', None)
        tr = get_object_or_404(TBLBANK, Bkid=Bkid)
        if tr.regdatetime:
            is_regist = True
        else:
            is_regist = False
        context = {'is_regist': is_regist}
    else:
        pass
    return HttpResponse(json.dumps(context), content_type="application/json")

@login_required(login_url='/')
def delete_row(request):
    context = {}
    page = request.POST.get('page', None)
    return HttpResponse(json.dumps(context), content_type="application/json")

@login_required(login_url='/')
def regist_division(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    if request.method == "POST":

        Bkid = request.POST.get('Bkid')
        tblbank_tr = TBLBANK.objects.get(Bkid=Bkid)
        Mid = request.user.username
        inoutType = request.POST.get('inoutType')
        itemId_list = request.POST.getlist('itemId_list')
        # print(len(itemId_list))
        subdivisionId_list = request.POST.getlist('subdivisionId_list')
        # print(len(subdivisionId_list))
        month = request.POST.get('month')
        Bkjukyo_list = request.POST.getlist('Bkjukyo_list')
        # print(len(Bkjukyo_list))
        Bkinout_list = request.POST.getlist('Bkinout_list')
        # print(len(Bkinout_list))
        Bkdate = request.POST.get('Bkdate')
        row = len(itemId_list) - 1

        try:
            close = Deadline.objects.get(business=business,year=Bkdate[:4],month=Bkdate[5:7])
            if close.regdatetime:
                return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
        except:
            pass

        sessionInfo = session_info(Bkdate[:4], Bkdate[5:7], business.session_month)
        start_date = datetime.datetime.strptime(Bkdate[:8]+'01', "%Y-%m-%d")
        a_month_ago = start_date - relativedelta(months=1)
        a_month_later = start_date + relativedelta(months=1)

        #--해당날짜 전월이월금 유무확인--
        try :
            premonth_transfer = Transaction.objects.filter(business=business, Bkdate=start_date).get(Bkdivision=0)
        except IndexError:
            return HttpResponse("<script>alert('IndexError');history.back();</script>")
        except Transaction.DoesNotExist:
            #--전월이월금 없는 경우 주계좌의 이전달 마지막 내역을 전월이월금으로 등록
            main = Account.objects.filter(business=business).get(main=True)
            data = TBLBANK.objects.filter(Bkacctno=main.account_number).filter(Bkdate__gte=a_month_ago).filter(Bkdate__lt=start_date).order_by('-Bkdate', '-Bkid').first()
            if data == None:
                return HttpResponse("<script>alert('주계좌의 전월잔액이 없습니다.');history.back();</script>")
            Transaction.objects.create(
                Bkid=data.Bkid,
                Bkdivision=0,
                Mid=Mid,
                business=business,
                Bkacctno=main.account_number,
                Bkname=main.bank.name,
                Bkdate=start_date,
                Bkjukyo="전월이월금",
                Bkinput=data.Bkjango,
                Bkoutput=0,
                Bkjango=data.Bkjango,
                item=Item.objects.get(
                    paragraph__subsection__year = sessionInfo['year'],
                    paragraph__subsection__institution=business.type3,
                    paragraph__subsection__code=0, paragraph__code=0, code=0),
                regdatetime=today
            )

        for r in range(row, 0, -1):
            latest_tr = Transaction.objects.filter(business=business, Bkdate__lte=Bkdate).order_by('-Bkdate','-id').first()
            jango = latest_tr.Bkjango
            Bkinput = 0
            Bkoutput = 0
            Bkjango = 0
            if inoutType == "input":
                Bkinput = Bkinout_list[r]
                Bkjango = jango + int(Bkinout_list[r])
            elif inoutType == "output":
                Bkoutput = Bkinout_list[r]
                Bkjango = jango - int(Bkinout_list[r])

            item = Item.objects.get(id=itemId_list[r])

            try:
                subdivision = Subdivision.objects.get(id=subdivisionId_list[r])
            except Subdivision.DoesNotExist:
                subdivision = None

            Transaction.objects.create(
                Bkid=Bkid,
                Bkdivision = r,
                Mid = Mid,
                Bkacctno = tblbank_tr.Bkacctno,
                Bkname = tblbank_tr.Bkname,
                Bkdate = Bkdate,
                Bkjukyo = "[분할]"+Bkjukyo_list[r],
                Bkinput = Bkinput,
                Bkoutput = Bkoutput,
                Bkjango = Bkjango,
                regdatetime = today,
                item = item,
                subdivision = subdivision,
                business = business,
            )

            update_list = Transaction.objects.filter(business=business, Bkdate__gt=Bkdate, Bkdate__lt=a_month_later)
            for update in update_list:
                if inoutType == "input":
                    update.Bkjango = int(update.Bkjango) + int(Bkinout_list[r])
                if inoutType == "output":
                    update.Bkjango = int(update.Bkjango) - int(Bkinout_list[r])
                update.save()

            tr, created = TBLBANK.objects.get_or_create(business=business, Bkid=Bkid, Bkdivision=r, Mid=Mid, Bkacctno=tblbank_tr.Bkacctno, Bkname=tblbank_tr.Bkname, Bkdate=tblbank_tr.Bkdate)
            tr.sub_Bkjukyo="[분할]"+Bkjukyo_list[r]
            tr.regdatetime=today
            tr.item = item
            tr.subdivision = subdivision
            tr.save()

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

'''
@login_required(login_url='/')
def extract_subdivision(request):
    #from django.core.serializers.json import DjangoJSONEncoder
    from django.core import serializers
    item_list = request.POST.get('item_list')
    print(item_list)
    #selected_item_list = []
    """
    subdivision_list = []
    for idx, val in enumerate(item_list):
        if val:
            selected_spi_list.append(int(val))
        else:
            selected_spi_list.append(0)
    """
    subdivisions = Subdivision.objects.filter(item__id = item_list)
    html_string="<option value="">--------------</option>"
    for sub in subdivisions:
        html_string += '<option value="%s">%s</option>' % (sub.id, sub.context)
    #subdivision_list.append(subdivisions)
    #subdivision_list = json.dumps(list(subdivisions), cls=DjangoJSONEncoder)
    subdivision_list = serializers.serialize('json', list(subdivisions))
    print(html_string)
    context = {'message': '성공', 'html_string': html_string}
    return HttpResponse(json.dumps(context), content_type="application/json")
'''

@login_required(login_url='/')
def popup_select_item(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    index = request.GET.get('index')
    row = request.GET.get('row')
    inoutType = request.GET.get('inoutType')
    if inoutType == "input":
        type_filter = "수입"
    elif inoutType == "output":
        type_filter = "지출"
    from django.db import connections
    query = """SELECT s.institution_id,s.code,s.context,s.type,p.code,p.context,i.code,i.context, i.id
    FROM accounting_subsection s
    LEFT OUTER JOIN accounting_paragraph p
    ON p.subsection_id = s.id
    LEFT OUTER JOIN accounting_item i
    ON i.paragraph_id = p.id
    WHERE s.institution_id = '"""+str(business.type3)+"""' and s.type = '"""+type_filter+"""' and s.code != 0
    ORDER BY s.institution_id, s.type, s.code, p.code, i.code"""
    with connections['default'].cursor() as cursor:
        cursor.execute(query)
        spi_list = cursor.fetchall()

    subdivisions = Subdivision.objects.filter(item__paragraph__subsection__institution=business.type3)
    return render(request, 'accounting/popup_select_item.html', {'spi_list': spi_list, 'subdivisions': subdivisions, 'index': index, 'row': row})

@login_required(login_url='/')
def annual_budget(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = int(DateFormat(today).format("Y"))
    if request.method == "POST":
        year = int(request.POST.get('year'))
    elif request.method == "GET":
        year = int(request.GET.get('year', this_year))

    filter_revenue = 'revenue'
    filter_expenditure = 'expenditure'
    if budget_type[:13] == "supplementary":
        filter_revenue = 'supplementary_revenue_'+budget_type[-1]
        filter_expenditure = 'supplementary_expenditure_'+budget_type[-1]

    total_revenue = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="수입", type=filter_revenue).aggregate(total=Coalesce(Sum('price'), 0))['total']
    total_expenditure = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="지출", type=filter_expenditure).aggregate(total=Coalesce(Sum('price'), 0))['total']
    total_difference = total_revenue - total_expenditure

    budget_list = []
    spi_list = []
    # spi_year = Item.objects.aggregate(year=Max('year'))
    sub_budget = []
    revenue_budget_page = ''
    expenditure_budget_page = ''
    supplementary_revenue_page = ''
    supplementary_expenditure_page = ''

    have_sub_bt = request.POST.get('have_sub_bt')

    # print("================ annual budget ====================")
    # print("business: ", business)
    # print("year: ", year)
    # print("have_sub_bt", have_sub_bt, type(have_sub_bt))
    # print("============== annual budget end ==================")

    if budget_type == "revenue" or budget_type[:21] == "supplementary_revenue":
        if have_sub_bt == "1":      # 본예산 불러오기
            budget_list = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="수입", type="revenue")
        elif have_sub_bt == "2":    # 전년도예산 불러오기
            print("전년도 세입 예산불러오기")
            budget_list = Budget.objects.filter(business=business, year=year-1, item__paragraph__subsection__type="수입", type="revenue")
            print(budget_list)
        else:
            budget_list = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="수입", type=budget_type)
        spi_list = item_info(year, business.type3, 'i')
        if budget_type == "revenue":
            revenue_budget_page = 'active'
        else:
            supplementary_revenue_page = 'active'
    elif budget_type == "expenditure" or budget_type[:25] == "supplementary_expenditure":
        if have_sub_bt == "1":
            budget_list = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="지출", type="expenditure")
        elif have_sub_bt == "2":
            print("전년도 세출 예산불러오기")
            budget_list = Budget.objects.filter(business=business, year=year-1, item__paragraph__subsection__type="지출", type="expenditure")
            print(budget_list)
        else:
            budget_list = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type="지출", type=budget_type)
        spi_list = item_info(year, business.type3, 'o')
        if budget_type == "expenditure":
            expenditure_budget_page = 'active'
        else:
            supplementary_expenditure_page = 'active'

    if budget_list:
        context_list = []
        unit_price_list = []
        cnt_list = []
        months_list = []
        percent_list = []
        sub_price_list = []
        for spi in spi_list:
            spi.budget_row = 1
            for index, budget in enumerate(budget_list):
                print(spi.code, budget.item.code)
                if spi.code == budget.item.code and spi.paragraph.code == budget.item.paragraph.code and spi.paragraph.subsection.code == budget.item.paragraph.subsection.code:
                    sub_columns = ['item','context','unit_price','cnt','months','percent','sub_price']
                    context_list = budget.context.split("|")
                    unit_price_list = budget.unit_price.split("|")
                    cnt_list = budget.cnt.split("|")
                    months_list = budget.months.split("|")
                    if budget.percent is not None:
                        percent_list = budget.percent.split("|")
                    sub_price_list = budget.sub_price.split("|")
                    row_list = []
                    for idx, val in enumerate(context_list):
                        r = []
                        if val != None:
                            r.append(budget.item.id)
                            r.append(context_list[idx])
                            r.append(unit_price_list[idx])
                            r.append(cnt_list[idx])
                            r.append(months_list[idx])
                            if budget.percent is not None:
                                r.append(percent_list[idx])
                            else:
                                r.append('')
                            r.append(sub_price_list[idx])
                        row_list.append(r)
                    sub_data = [ dict(zip(sub_columns,row)) for row in row_list ]
                    spi.budget_price = budget.price
                    spi.budget_row = budget.row
                    spi.sub_budget = sub_data
                    print(sub_data)
                    sub_budget += sub_data

    return render(request, 'accounting/annual_budget.html', {'total_revenue': total_revenue, 'total_expenditure': total_expenditure, 'total_difference': total_difference,'budget_type': budget_type, 'sub_budget': sub_budget, 'spi_list': spi_list, 'budget_list': budget_list, 'budget_management': 'active', 'revenue_budget_page': revenue_budget_page, 'expenditure_budget_page': expenditure_budget_page, 'supplementary_revenue_page': supplementary_revenue_page, 'supplementary_expenditure_page': supplementary_expenditure_page, 'master_login': request.session['master_login'], 'business': business, 'year_range': range(this_year+1, 1999, -1), 'year': year})

@login_required(login_url='/')
def print_yearly_budget(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "POST":
        year = int(request.POST.get('year'))

    if budget_type in ['revenue', 'supplementary_revenue']:
        stype_filter = '수입'
        temp_budget_type = 'revenue'    # 본예산 조회 시 전년도 예산구분 확인 시 사용
    elif budget_type in ['expenditure', 'supplementary_expenditure']:
        stype_filter='지출'
        temp_budget_type = 'expenditure'    # 본예산 조회 시 전년도 예산구분 확인 시 사용

    # 추경 및 본예산 등록확인
    filter_budget_type = getLatestBudgetType(business, year, budget_type)
    if filter_budget_type is None:  # 등록 안 된 경우 알림 및 창닫기
        if 'supplementary' in budget_type:
            return HttpResponse("<script>alert('추경예산이 등록되지 않았습니다.');close();</script>")
        else:
            return HttpResponse("<script>alert('본예산이 등록되지 않았습니다.');close();</script>")
    # 전년도 최근예산 타입조회 (본예산 조회 시 사용)
    x_budget_type = getLatestBudgetType(business, year - 1, temp_budget_type)

    total = 0
    xtotal = 0
    dtotal = 0

    subsection_list = Subsection.objects.filter(year=year, type=stype_filter, institution=business.type3).exclude(code=0)
    for subsection in subsection_list:
        s_total = 0
        xs_total = 0

        paragraph_list = Paragraph.objects.filter(subsection=subsection)
        for paragraph in paragraph_list:
            p_total = 0
            xp_total = 0

            item_list = Item.objects.filter(paragraph=paragraph)
            for item in item_list:
                try:
                    if 'supplementary' in budget_type:
                        budget = Budget.objects.get(business=business, year=year, item=item, type=filter_budget_type)
                        item.xi_total = Budget.objects.get(business=business, year=year, item=item, type=budget_type[14:]).price
                    else:
                        budget = Budget.objects.get(business=business, year=year, item=item, type=budget_type)
                        try:
                            # 전년도 예산액 : 전년도 추경예산 있는 경우 추경예산, 없는 경우 본예산 출력
                            item.xi_total = Budget.objects.get(
                                business=business, year=year - 1, item__paragraph__subsection__code=subsection.code,
                                item__paragraph__code=paragraph.code, item__code=item.code, type=x_budget_type).price
                        except:
                            item.xi_total = 0
                    item.i_total = budget.price

                    percent_list = []
                    sub_columns = ['item','context','unit_price','cnt','months','percent','sub_price']
                    context_list = budget.context.split("|")
                    unit_price_list = budget.unit_price.split("|")
                    cnt_list = budget.cnt.split("|")
                    months_list = budget.months.split("|")
                    if budget.percent is not None:
                        percent_list = budget.percent.split("|")
                    sub_price_list = budget.sub_price.split("|")
                    row_list = []
                    for idx, val in enumerate(context_list):
                        r = []
                        if val != None:
                            r.append(budget.item.id)
                            r.append(context_list[idx])
                            r.append(unit_price_list[idx])
                            r.append(cnt_list[idx])
                            r.append(months_list[idx])
                            if budget.percent is not None:
                                r.append(percent_list[idx])
                            else:
                                r.append('')
                            r.append(sub_price_list[idx])
                        row_list.append(r)
                        item.sub_data = [ dict(zip(sub_columns,row)) for row in row_list ]
                except Exception as ex:
                    print(ex)
                    item.i_total = 0
                    print("except i_total:", item.i_total)
                    if 'supplementary' in budget_type:
                        item.xi_total = Budget.objects.get(business=business, year=year, item=item, type=budget_type[14:]).price
                    else:
                        item.xi_total = Budget.objects.get(
                            business=business, year=year - 1, item__paragraph__subsection__code=subsection.code,
                            item__paragraph__code=paragraph.code, item__code=item.code, type=x_budget_type).price
                    item.sub_data = []

                item.di_total = item.i_total - item.xi_total
                p_total += item.i_total
                xp_total += item.xi_total

            paragraph.item_list = item_list
            paragraph.p_total = p_total
            paragraph.xp_total = xp_total
            paragraph.dp_total = paragraph.p_total - paragraph.xp_total
            s_total += paragraph.p_total
            xs_total += paragraph.xp_total

        subsection.paragraph_list = paragraph_list
        subsection.s_total = s_total
        subsection.xs_total = xs_total
        subsection.ds_total = subsection.s_total - subsection.xs_total
        total += subsection.s_total
        xtotal += subsection.xs_total
        dtotal += subsection.ds_total

    page_list = []
    sub_row = []
    for index, subsection in enumerate(subsection_list):
        # print(index, subsection)
        sub_row.append(subsection)
        if (budget_type in ['revenue', 'supplementary_revenue', 'supplementary_revenue']) and (index+1 in [3,9]):
            page_list.append(sub_row)
            # print(index, sub_row)
            sub_row = []
        elif (budget_type in ['expenditure', 'supplementary_expenditure', 'supplementary_expenditure']) and (index+1 in [1,2,5,10]):
            page_list.append(sub_row)
            # print(index, sub_row)
            sub_row = []

    return render(request, 'accounting/print_yearly_budget.html', {
        'budget_type': budget_type, 'year': year, 'page_list': page_list,
        'total': total, 'xtotal': xtotal, 'dtotal': dtotal
        })

@login_required(login_url='/')
def print_supplementary_budget(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "POST":
        year = int(request.POST.get('year'))

    if budget_type in ['revenue', 'supplementary_revenue']:
        stype_filter = '수입'
        degree = int(request.POST.get('revenue_degree'))
    elif budget_type in ['expenditure', 'supplementary_expenditure']:
        stype_filter='지출'
        degree = int(request.POST.get('expenditure_degree'))

    # 추경예산 등록확인
    if not validBudgetDegree(business, year, degree):  # 추경예산 안 된 경우 알림 및 창닫기
        return HttpResponse("<script>alert('추경예산이 등록되지 않았습니다.');close();</script>")

    # 비교대상 설정 : 이전예산 (1차추경은 본예산(0), n차 추경은 n-1차 추경)
    x_degree = degree - 1

    total = 0
    xtotal = 0
    dtotal = 0

    subsection_list = Subsection.objects.filter(year=year, type=stype_filter, institution=business.type3).exclude(code=0)
    for subsection in subsection_list:
        s_total = 0
        xs_total = 0

        paragraph_list = Paragraph.objects.filter(subsection=subsection)
        for paragraph in paragraph_list:
            p_total = 0
            xp_total = 0

            item_list = Item.objects.filter(paragraph=paragraph)
            for item in item_list:
                # 선택한 추경예산정보 불러오기
                try:
                    budget = Budget.objects.get(business=business, year=year, item=item, degree=degree)
                    item.i_total = budget.price

                    percent_list = []
                    sub_columns = ['item','context','unit_price','cnt','months','percent','sub_price']
                    context_list = budget.context.split("|")
                    unit_price_list = budget.unit_price.split("|")
                    cnt_list = budget.cnt.split("|")
                    months_list = budget.months.split("|")
                    if budget.percent is not None:
                        percent_list = budget.percent.split("|")
                    sub_price_list = budget.sub_price.split("|")
                    row_list = []
                    for idx, val in enumerate(context_list):
                        r = []
                        if val != None:
                            r.append(budget.item.id)
                            r.append(context_list[idx])
                            r.append(unit_price_list[idx])
                            r.append(cnt_list[idx])
                            r.append(months_list[idx])
                            if budget.percent is not None:
                                r.append(percent_list[idx])
                            else:
                                r.append('')
                            r.append(sub_price_list[idx])
                        row_list.append(r)
                        item.sub_data = [ dict(zip(sub_columns,row)) for row in row_list ]
                except Exception as ex:
                    print(ex)
                    item.i_total = 0
                    item.sub_data = []

                # 선택한 추경예산의 이전예산정보 불러오기
                try :
                    item.xi_total = Budget.objects.get(business=business, year=year, item=item, degree=x_degree).price
                except Exception as ex:
                    print(ex)
                    item.xi_total = 0

                item.di_total = item.i_total - item.xi_total
                p_total += item.i_total
                xp_total += item.xi_total

            paragraph.item_list = item_list
            paragraph.p_total = p_total
            paragraph.xp_total = xp_total
            paragraph.dp_total = paragraph.p_total - paragraph.xp_total
            s_total += paragraph.p_total
            xs_total += paragraph.xp_total

        subsection.paragraph_list = paragraph_list
        subsection.s_total = s_total
        subsection.xs_total = xs_total
        subsection.ds_total = subsection.s_total - subsection.xs_total
        total += subsection.s_total
        xtotal += subsection.xs_total
        dtotal += subsection.ds_total

    page_list = []
    sub_row = []
    for index, subsection in enumerate(subsection_list):
        # print(index, subsection)
        sub_row.append(subsection)
        if (budget_type in ['revenue', 'supplementary_revenue', 'supplementary_revenue']) and (index+1 in [3,9]):
            page_list.append(sub_row)
            # print(index, sub_row)
            sub_row = []
        elif (budget_type in ['expenditure', 'supplementary_expenditure', 'supplementary_expenditure']) and (index+1 in [1,2,5,10]):
            page_list.append(sub_row)
            # print(index, sub_row)
            sub_row = []

    return render(request, 'accounting/print_supplementary_budget.html', {
        'budget_type': budget_type, 'degree': degree, 'year': year, 'page_list': page_list,
        'total': total, 'xtotal': xtotal, 'dtotal': dtotal
        })

@login_required(login_url='/')
def regist_annual_budget(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "POST":
        budget_type = request.POST.get('budget_type')
        budget_year = request.POST.get('budget_year')
        budget_spi = request.POST.getlist('budget_spi')
        budget_row = request.POST.getlist('budget_row')
        budget_price = request.POST.getlist('budget_price')
        budget_context = request.POST.getlist('budget_context')
        budget_unit_price = request.POST.getlist('budget_unit_price')
        budget_cnt = request.POST.getlist('budget_cnt')
        budget_months = request.POST.getlist('budget_months')
        budget_percent = request.POST.getlist('budget_percent')
        budget_sub_price = request.POST.getlist('budget_sub_price')
        degree = budget_type[-1] if 'supplementary' in budget_type else 0

        row_idx = 1
        for idx, val in enumerate(budget_spi):
            if idx:
                context_list = ''
                unit_price_list = ''
                cnt_list = ''
                months_list = ''
                percent_list = ''
                sub_price_list = ''
                for row in range(1, int(budget_row[idx])+1):
                    context_list += budget_context[row_idx]+'|'
                    unit_price_list += budget_unit_price[row_idx]+'|'
                    cnt_list += budget_cnt[row_idx]+'|'
                    months_list += budget_months[row_idx]+'|'
                    percent_list += budget_percent[row_idx]+'|'
                    sub_price_list += budget_sub_price[row_idx]+'|'
                    row_idx += 1
                context_list = context_list[:-1]
                unit_price_list = unit_price_list[:-1]
                cnt_list = cnt_list[:-1]
                months_list = months_list[:-1]
                percent_list = percent_list[:-1]
                sub_price_list = sub_price_list[:-1]

                #이미 등록되어있는 경우 새로등록이 아닌 찾아 바꾸기
                budget, created = Budget.objects.get_or_create(business=business, year=budget_year, item=Item.objects.get(id=budget_spi[idx]), type=budget_type, degree=degree, defaults={'row': int(budget_row[idx]), 'price': 0, 'context': '', 'unit_price': '', 'cnt': '', 'months': '', 'percent': '', 'sub_price': '0'})
                budget.row = int(budget_row[idx])
                if budget_price[idx] != '':
                    budget.price = int(budget_price[idx])
                budget.context = context_list
                budget.unit_price = unit_price_list
                budget.cnt = cnt_list
                budget.months = months_list
                budget.percent = percent_list
                budget.sub_price = sub_price_list
                budget.save()

    response = redirect('annual_budget', budget_type)
    response['Location'] += '?year='+budget_year
    return response

@login_required(login_url='/')
def budget_settlement(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    s_type = request.GET.get('s_type', '1') #search_type
    year = int(request.GET.get('year', DateFormat(today).format("Y")))
    month = int(request.GET.get('month', business.session_month))

    sessionInfo = session_info(str(year), str(month), business.session_month)


    if s_type == '1':
        start_date = datetime.datetime.strptime(str(year)+'-'+business.session_month+'-01', '%Y-%m-%d')
        print(start_date)
        end_date = datetime.datetime.strptime(str(year+1)+'-'+business.session_month+'-01', '%Y-%m-%d')
        print(end_date)
    else:
        start_date = datetime.datetime.strptime(str(year)+'-'+business.session_month+'-01', '%Y-%m-%d')
        end_date = start_date + relativedelta(months=1)

    print(start_date, end_date)

    revenue_settlement_page = ''
    expenditure_settlement_page = ''

    total_budget = 0
    now_budget = 0
    total_sum = 0
    total_difference = 0

    if budget_type == "revenue":
        filter_type = "수입"
    elif budget_type == "expenditure":
        filter_type = "지출"

    filter_budget_type = getLatestBudgetType(business, year, budget_type)

    subsection_list = Subsection.objects.filter(year=sessionInfo['year'], institution=business.type3, type=filter_type).annotate(count=Count('paragraph__item')).exclude(count=0).exclude(code=0)
    paragraph_list = Paragraph.objects.filter(subsection__year=sessionInfo['year'], subsection__institution=business.type3, subsection__type=filter_type).annotate(count=Count('item')).exclude(count=0).exclude(code=0)
    item_list = Item.objects.filter(
        paragraph__subsection__year=sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type = filter_type
    ).annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = year, then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0)).exclude(code=0)
    if budget_type == "revenue":
        item_list2 = Item.objects.filter(
                paragraph__subsection__year = sessionInfo['year'],
                paragraph__subsection__institution = business.type3,
                paragraph__subsection__type = filter_type
        ).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkinput'))))))), 0)
        ).exclude(code=0)
        item_list3 = Item.objects.filter(
                paragraph__subsection__year = sessionInfo['year'],
                paragraph__subsection__institution = business.type3,
                paragraph__subsection__type = filter_type
        ).annotate(
            now_budget=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = sessionInfo['start_date'], then=Case(
                When(transaction__Bkdate__lt = start_date, then='transaction__Bkinput'))))))), 0)
        ).exclude(code=0)
        revenue_settlement_page = 'active'

    elif budget_type == "expenditure":
        item_list2 = Item.objects.filter(
                paragraph__subsection__year = sessionInfo['year'],
                paragraph__subsection__institution = business.type3,
                paragraph__subsection__type = filter_type
        ).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkoutput'))))))), 0)
        ).exclude(code=0)
        item_list3 = Item.objects.filter(
                paragraph__subsection__year = sessionInfo['year'],
                paragraph__subsection__institution = business.type3,
                paragraph__subsection__type = filter_type
        ).annotate(
            now_budget=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = sessionInfo['start_date'], then=Case(
                When(transaction__Bkdate__lt = start_date, then='transaction__Bkoutput'))))))), 0)
        ).exclude(code=0)
        expenditure_settlement_page = 'active'

    for idx, val in enumerate(item_list):
        item_list[idx].total_sum = item_list2[idx].total_sum
        item_list[idx].now_budget = item_list[idx].total_budget - item_list3[idx].now_budget
        item_list[idx].total_difference = item_list[idx].now_budget - item_list[idx].total_sum
        print(item_list[idx].now_budget, item_list[idx].total_difference)
        total_budget += item_list[idx].total_budget
        now_budget += item_list[idx].now_budget
        total_sum += item_list[idx].total_sum
        total_difference += item_list[idx].total_difference

    return render(request,'accounting/budget_settlement.html', {
        'settlement_management': 'active', 'master_login': request.session['master_login'],
        'revenue_settlement_page': revenue_settlement_page, 'expenditure_settlement_page': expenditure_settlement_page,
        'business': business, 'year_range': range(today.year, 1999, -1), 'year': year,
        'month_range': range(1, 13), 'month': month, 's_type': s_type,
        'subsection_list': subsection_list, 'paragraph_list': paragraph_list,
        'item_list': item_list, 'budget_type': budget_type,
        'total_budget': total_budget, 'now_budget': now_budget,
        'total_sum': total_sum, 'total_difference': total_difference})

@login_required(login_url='/')
def trial_balance(request):
    if request.method == "GET":
        business = get_object_or_404(Business, pk=request.session['business'])
        today = datetime.datetime.now()
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        year2 = int(request.GET.get('year2', today.year))
        month2 = int(request.GET.get('month2', today.month))

        sessionInfo = session_info(str(year), business.session_month, business.session_month)

        start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
        if month2 <= 11:
            end_date = datetime.datetime.strptime(str(year2)+'-'+str(month2+1)+'-01', '%Y-%m-%d')
        else:
            end_date = datetime.datetime.strptime(str(year2+1)+'-'+str(month2-11)+'-01', '%Y-%m-%d')

        item_list = Item.objects.filter(
            paragraph__subsection__year=sessionInfo['year'],
            paragraph__subsection__institution=business.type3
        ).order_by('paragraph__subsection__type', 'paragraph__subsection__code', 'paragraph__code', 'code'
        ).exclude(code=0)

        item_list.cumulative_income_total = 0  # 수입누계 총합
        item_list.income_total = 0  # 수입금액 총합
        item_list.expenditure_total = 0  # 지출금액 총합
        item_list.cumulative_expenditure_total = 0  # 지출누계 총합
        for idx, item in enumerate(item_list):
            if item.paragraph.subsection.type == '수입':
                # 목별 수입누계
                item_list[idx].cumulative_income = getTransactionSumByItem(
                    business, item, sessionInfo['start_date'], end_date)
                item_list.cumulative_income_total += item_list[idx].cumulative_income

                # 목별 수입금액
                item_list[idx].income = getTransactionSumByItem(
                    business, item, start_date, end_date)
                item_list.income_total += item_list[idx].income
            elif item.paragraph.subsection.type == '지출':
                # 목별 지출금액
                item_list[idx].expenditure = getTransactionSumByItem(
                    business, item, start_date, end_date)
                item_list.expenditure_total += item_list[idx].expenditure

                # 목별 수입누계
                item_list[idx].cumulative_expenditure = getTransactionSumByItem(
                    business, item, sessionInfo['start_date'], end_date)
                item_list.cumulative_expenditure_total += item_list[idx].cumulative_expenditure

    return render(request, 'accounting/trial_balance.html', {
        'settlement_management': 'active', 'trial_balance_page': 'active',
        'year_range': range(today.year+1, 1999, -1), 'month_range': range(1, 13),
        'year': year, 'month': month, 'year2': year2, 'month2': month2,
        'item_list': item_list
    })

@login_required(login_url='/')
def annual_trial_balance(request):
    if request.method == "GET":
        business = get_object_or_404(Business, pk=request.session['business'])
        today = datetime.datetime.now()
        cyear = int(DateFormat(today).format("Y"))
        year = int(request.GET.get('year', cyear))

        sessionInfo = session_info(str(year), business.session_month, business.session_month)

        # 당해년도 목(Item) 목록
        item_list = Item.objects.filter(
            paragraph__subsection__year=sessionInfo['year'],
            paragraph__subsection__institution=business.type3
        ).order_by('paragraph__subsection__type', 'paragraph__subsection__code', 'paragraph__code', 'code').exclude(
            code=0)

        ym_list = []
        for x in range(0, 12):
            date = sessionInfo['start_date'] + relativedelta(months=x)
            ym_list.append({'y': DateFormat(date).format("Y"), 'm': DateFormat(date).format("m")})

        input_budget_total = 0  # 세입예산합계
        output_budget_total = 0  # 세출예산합계
        input_settlement_total = 0  # 세입결산합계
        output_settlement_total = 0  # 세출결산합계
        revenue_budget_type = getLatestBudgetType(business, year, "revenue")
        expenditure_budget_type = getLatestBudgetType(business, year, "expenditure")
        for idx, item in enumerate(item_list):
            # 예산액 (본예산을 기준으로 함)
            try:
                filter_budget_type = getLatestBudgetType(business, year, "both")
                item_list[idx].budget_amount = Budget.objects.filter(
                    Q(type=revenue_budget_type) | Q(type=expenditure_budget_type)
                ).get(business=business, year=year, item=item.pk).price
                if item.paragraph.subsection.type == "수입":
                    input_budget_total += item_list[idx].budget_amount
                elif item.paragraph.subsection.type == "지출":
                    output_budget_total += item_list[idx].budget_amount
            except:
                item_list[idx].budget_amount = 0

            # 결산액
            settlement = Transaction.objects.filter(
                business=business, Bkdate__gte=sessionInfo['start_date'],
                Bkdate__lt=sessionInfo['end_date'], item=item.pk
            ).aggregate(
                input=Coalesce(Sum('Bkinput'), 0),
                output=Coalesce(Sum('Bkoutput'), 0)
            )
            if item.paragraph.subsection.type == "수입":
                item_list[idx].settlement_amount = settlement['input']
                input_settlement_total += item_list[idx].settlement_amount
            elif item.paragraph.subsection.type == "지출":
                item_list[idx].settlement_amount = settlement['output']
                output_settlement_total += item_list[idx].settlement_amount

            # 월별 시산액
            ms_list = []
            for ym in ym_list:
                start_date = datetime.datetime.strptime(ym['y'] + '-' + ym['m'] + '-01', '%Y-%m-%d')
                end_date = start_date + relativedelta(months=1)

                settlement = Transaction.objects.filter(
                    business=business, Bkdate__gte=start_date,
                    Bkdate__lt=end_date, item=item.pk
                ).aggregate(
                    input=Coalesce(Sum('Bkinput'), 0),
                    output=Coalesce(Sum('Bkoutput'), 0)
                )
                if item.paragraph.subsection.type == "수입":
                    ms_list.append(settlement['input'])
                elif item.paragraph.subsection.type == "지출":
                    ms_list.append(settlement['output'])
                else:
                    ms_list.append(0)
            item_list[idx].ms_list = ms_list

        item_list.input_budget_total = input_budget_total
        item_list.output_budget_total = output_budget_total
        item_list.input_settlement_total = input_settlement_total
        item_list.output_settlement_total = output_settlement_total

        # 월별 시산액 합계
        month_total_list = []
        for ym in ym_list:
            start_date = datetime.datetime.strptime(ym['y'] + '-' + ym['m'] + '-01', '%Y-%m-%d')
            end_date = start_date + relativedelta(months=1)

            # 세입 월별시산액 합계 (code=0은 제외)
            input_month_total = Transaction.objects.filter(
                business=business, item__paragraph__subsection__year=year,
                Bkdate__gte=start_date, Bkdate__lt=end_date,
                item__paragraph__subsection__type="수입"
            ).exclude(item__code=0).aggregate(input=Coalesce(Sum('Bkinput'), 0))['input']
            output_month_total = Transaction.objects.filter(
                business=business, item__paragraph__subsection__year=year,
                Bkdate__gte=start_date, Bkdate__lt=end_date,
                item__paragraph__subsection__type="지출"
            ).exclude(item__code=0).aggregate(output=Coalesce(Sum('Bkoutput'), 0))['output']
            month_total_list.append({'input': input_month_total, 'output': output_month_total})

    return render(request, 'accounting/annual_trial_balance.html', {
        'settlement_management': 'active', 'annual_trial_balance_page': 'active',
        'ym_list': ym_list, 'year_range': range(cyear+1, 1999, -1),'year': year,
        'item_list': item_list, 'month_total_list': month_total_list
    })

@login_required(login_url='/')
def print_annual_trial_balance(request):
    if request.method == "POST":
        business = get_object_or_404(Business, pk=request.session['business'])
        year = int(request.POST.get('year'))

        sessionInfo = session_info(str(year), business.session_month, business.session_month)

        # 당해년도 목(Item) 목록
        item_list = Item.objects.filter(
            paragraph__subsection__year = sessionInfo['year'],
            paragraph__subsection__institution = business.type3
        ).order_by('paragraph__subsection__type', 'paragraph__subsection__code', 'paragraph__code', 'code').exclude(code=0)

        ym_list = []
        for x in range(0, 12):
            date = sessionInfo['start_date'] + relativedelta(months=x)
            ym_list.append({'y': DateFormat(date).format("Y"), 'm': DateFormat(date).format("m")})

        input_budget_total = 0      # 세입예산합계
        output_budget_total = 0     # 세출예산합계
        input_settlement_total = 0  # 세입결산합계
        output_settlement_total = 0 # 세출결산합계
        revenue_budget_type = getLatestBudgetType(business, year, "revenue")
        expenditure_budget_type = getLatestBudgetType(business, year, "expenditure")
        for idx, item in enumerate(item_list):
            # 예산액 (본예산을 기준으로 함)
            try :
                item_list[idx].budget_amount = Budget.objects.filter(
                    Q(type=revenue_budget_type)|Q(type=expenditure_budget_type)
                ).get(business=business, year=year, item = item.pk).price
                if item.paragraph.subsection.type == "수입" :
                    input_budget_total += item_list[idx].budget_amount
                elif item.paragraph.subsection.type == "지출" :
                    output_budget_total += item_list[idx].budget_amount
            except :
                item_list[idx].budget_amount = 0

            # 결산액
            settlement = Transaction.objects.filter(
                business = business, Bkdate__gte = sessionInfo['start_date'],
                Bkdate__lt = sessionInfo['end_date'], item = item.pk
            ).aggregate(
                input=Coalesce(Sum('Bkinput'),0),
                output=Coalesce(Sum('Bkoutput'),0)
            )
            if item.paragraph.subsection.type == "수입" :
                item_list[idx].settlement_amount = settlement['input']
                input_settlement_total += item_list[idx].settlement_amount
            elif item.paragraph.subsection.type == "지출" :
                item_list[idx].settlement_amount = settlement['output']
                output_settlement_total += item_list[idx].settlement_amount

            # 월별 시산액
            ms_list = []
            for ym in ym_list:
                start_date = datetime.datetime.strptime(ym['y']+'-'+ym['m']+'-01', '%Y-%m-%d')
                end_date = start_date + relativedelta(months=1)

                settlement = Transaction.objects.filter(
                    business = business, Bkdate__gte = start_date,
                    Bkdate__lt = end_date, item = item.pk
                ).aggregate(
                    input=Coalesce(Sum('Bkinput'),0),
                    output=Coalesce(Sum('Bkoutput'),0)
                )
                if item.paragraph.subsection.type == "수입" :
                    ms_list.append(settlement['input'])
                elif item.paragraph.subsection.type == "지출" :
                    ms_list.append(settlement['output'])
                else :
                    ms_list.append(0)
            item_list[idx].ms_list = ms_list

        item_list.input_budget_total = input_budget_total
        item_list.output_budget_total = output_budget_total
        item_list.input_settlement_total = input_settlement_total
        item_list.output_settlement_total = output_settlement_total

        # 월별 시산액 합계
        month_total_list = []
        for ym in ym_list:
            start_date = datetime.datetime.strptime(ym['y']+'-'+ym['m']+'-01', '%Y-%m-%d')
            end_date = start_date + relativedelta(months=1)

            # 세입 월별시산액 합계 (code=0은 제외)
            input_month_total = Transaction.objects.filter(
                business = business, item__paragraph__subsection__year = year,
                Bkdate__gte = start_date, Bkdate__lt = end_date,
                item__paragraph__subsection__type = "수입"
            ).exclude(item__code=0).aggregate(input=Coalesce(Sum('Bkinput'),0))['input']
            output_month_total = Transaction.objects.filter(
                business = business, item__paragraph__subsection__year = year,
                Bkdate__gte = start_date, Bkdate__lt = end_date,
                item__paragraph__subsection__type = "지출"
            ).exclude(item__code=0).aggregate(output=Coalesce(Sum('Bkoutput'),0))['output']
            month_total_list.append({'input': input_month_total, 'output': output_month_total})

    return render(request, 'accounting/print_annual_trial_balance.html', {
        'settlement_management': 'active', 'annual_trial_balance_page': 'active',
        'ym_list': ym_list, 'year': year, 'item_list': item_list,
        'month_total_list': month_total_list
    })

@login_required(login_url='/')
def print_settlement(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = int(DateFormat(today).format("Y"))
    this_month = int(DateFormat(today).format("m"))
    year = this_year
    month = 1
    year2 = this_year
    month2 = this_month
    if business.type3_id == "어린이집":
        month = 3
        if this_month < 3:
            year = this_year - 1
    return render(request,'accounting/print_settlement.html', {'print_menu': 'active', 'print_settlement_page': 'active', 'master_login': request.session['master_login'], 'business': business, 'year_range': range(this_year, 1999, -1), 'month_range': range(1, 13), 'year': year, 'month': month, 'year2': year2, 'month2': month2})

@login_required(login_url='/')
def print_budget(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = int(DateFormat(today).format("Y"))
    this_month = int(DateFormat(today).format("m"))
    year = this_year
    month = 1
    year2 = this_year
    month2 = this_month
    if business.type3_id == "어린이집":
        month = 3
        if this_month < 3:
            year = this_year - 1
    return render(request,'accounting/print_budget.html', {'print_menu': 'active', 'print_budget_page': 'active', 'master_login': request.session['master_login'], 'business': business, 'year_range': range(this_year, 1999, -1), 'month_range': range(1, 13), 'year': year, 'month': month, 'year2': year2, 'month2': month2})

@login_required(login_url='/')
def print_budget_settlement(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    start_date = datetime.datetime.strptime(str(year)+'-03-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year+1)+'-03-01', '%Y-%m-%d')

    total_budget = 0
    total_sum = 0
    total_difference = 0

    if budget_type == "revenue":
        filter_type = "수입"
    elif budget_type == "expenditure":
        filter_type = "지출"

    filter_budget_type = getLatestBudgetType(business, year, budget_type)

    subsection_list = Subsection.objects.filter(year=year, institution=business.type3, type=filter_type).annotate(count=Count('paragraph__item')).exclude(count=0).exclude(code=0)
    paragraph_list = Paragraph.objects.filter(subsection__year=year, subsection__institution=business.type3, subsection__type=filter_type).annotate(count=Count('item')).exclude(count=0)
    item_list = Item.objects.filter(
        paragraph__subsection__year = year,
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type = filter_type
    ).exclude(code=0).annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = year, then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0))
    if budget_type == "revenue":
        item_list2 = Item.objects.filter(
            paragraph__subsection__year = year,
            paragraph__subsection__institution = business.type3,
            paragraph__subsection__type = filter_type
        ).exclude(code=0).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkinput'))))))), 0))
    elif budget_type == "expenditure":
        item_list2 = Item.objects.filter(
            paragraph__subsection__year = year,
            paragraph__subsection__institution = business.type3,
            paragraph__subsection__type = filter_type
        ).exclude(code=0).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkoutput'))))))), 0))

    for idx, val in enumerate(item_list):
        item_list[idx].total_sum = item_list2[idx].total_sum
        item_list[idx].total_difference = item_list[idx].total_budget-item_list[idx].total_sum
        total_budget += item_list[idx].total_budget
        total_sum += item_list[idx].total_sum
        total_difference += item_list[idx].total_difference

    return render(request,'accounting/print_budget_settlement.html', {
        'settlement_management': 'active', 'master_login': request.session['master_login'],
        'business': business, 'year': year, 'subsection_list': subsection_list,
        'paragraph_list': paragraph_list, 'item_list': item_list,
        'total_budget': total_budget, 'total_sum': total_sum,
        'total_difference': total_difference, 'budget_type': budget_type})

@login_required(login_url='/')
def print_budget_settlement2(request, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    year2 = int(request.POST.get('year2', year))
    month2 = int(request.POST.get('month2', month))

    sessionInfo = session_info(str(year), str(month), business.session_month)

    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year2)+'-'+str(month2)+'-01', '%Y-%m-%d') + relativedelta(months=1)

    total_budget = 0
    now_budget = 0
    total_sum = 0
    total_difference = 0

    if budget_type == "revenue":
        filter_type = "수입"
    elif budget_type == "expenditure":
        filter_type = "지출"

    filter_budget_type = getLatestBudgetType(business, year, budget_type)

    subsection_list = Subsection.objects.filter(year=sessionInfo['year'], institution=business.type3, type=filter_type).annotate(count=Count('paragraph__item')).exclude(count=0).exclude(code=0)
    paragraph_list = Paragraph.objects.filter(subsection__year=sessionInfo['year'], subsection__institution=business.type3, subsection__type=filter_type).annotate(count=Count('item')).exclude(count=0)
    item_list = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type = filter_type
    ).exclude(code=0).annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = sessionInfo['year'], then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0)).exclude(code=0)
    if budget_type == "revenue":
        item_list2 = Item.objects.filter(
            paragraph__subsection__year = sessionInfo['year'],
            paragraph__subsection__institution = business.type3,
            paragraph__subsection__type = filter_type
        ).exclude(code=0).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkinput'))))))), 0))
        item_list3 = Item.objects.filter(
            paragraph__subsection__year = sessionInfo['year'],
            paragraph__subsection__institution = business.type3,
            paragraph__subsection__type = filter_type
        ).annotate(
            now_budget=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = sessionInfo['start_date'], then=Case(
                When(transaction__Bkdate__lt = start_date, then='transaction__Bkinput'))))))), 0))
    elif budget_type == "expenditure":
        item_list2 = Item.objects.filter(
            paragraph__subsection__year = sessionInfo['year'],
            paragraph__subsection__institution = business.type3,
            paragraph__subsection__type = filter_type
        ).exclude(code=0).annotate(
            total_sum=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = start_date, then=Case(
                When(transaction__Bkdate__lt = end_date, then='transaction__Bkoutput'))))))), 0)).exclude(code=0)
        item_list3 = Item.objects.filter(
            paragraph__subsection__year = sessionInfo['year'],
            paragraph__subsection__institution = business.type3,
            paragraph__subsection__type = filter_type
        ).annotate(
            now_budget=Coalesce(Sum(Case(
                When(transaction__business = business, then=Case(
                When(transaction__Bkdate__gte = sessionInfo['start_date'], then=Case(
                When(transaction__Bkdate__lt = start_date, then='transaction__Bkoutput'))))))), 0)).exclude(code=0)

    for idx, val in enumerate(item_list):
        item_list[idx].now_budget = item_list[idx].total_budget - item_list3[idx].now_budget
        item_list[idx].total_sum = item_list2[idx].total_sum
        item_list[idx].total_difference = item_list[idx].now_budget-item_list[idx].total_sum
        total_budget += item_list[idx].total_budget
        now_budget += item_list[idx].now_budget
        total_sum += item_list[idx].total_sum
        total_difference += item_list[idx].total_difference

    return render(request,'accounting/print_budget_settlement2.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'subsection_list': subsection_list, 'paragraph_list': paragraph_list, 'item_list': item_list, 'total_budget': total_budget, 'total_sum': total_sum, 'total_difference': total_difference, 'budget_type': budget_type, 'now_budget': now_budget})

@login_required(login_url='/')
def print_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    year2 = int(request.POST.get('year2', year))
    month2 = int(request.POST.get('month2', month))

    sessionInfo = session_info(str(year), str(month), business.session_month)

    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year2)+'-'+str(month2)+'-01', '%Y-%m-%d') + relativedelta(months=1)

    transaction_list = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date).order_by('Bkdate','Bkid','Bkdivision').exclude(item=None)

    ym_list = []
    for transaction in transaction_list:
        ym_list.append(str(transaction.Bkdate)[:7])
    ym_list = list(set(ym_list))
    ym_list.sort()

    ym_range = []
    total_input = 0
    total_output = 0
    for ym in ym_list:
        transaction = []
        ym_tr = Transaction.objects.filter(business = business, Bkdate__year=ym[:4], Bkdate__month=ym[-2:]).order_by('Bkdate','id')
        tr_paginator = Paginator(ym_tr, 40)
        for tr_page in range(1, tr_paginator.num_pages+1):
            transaction.append(tr_paginator.page(tr_page))

        total = Transaction.objects.filter(business=business, Bkdate__year=ym[:4], Bkdate__month=ym[-2:]).aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
        sub_date =  datetime.datetime.strptime(ym[:4]+'-'+ym[-2:]+'-01', '%Y-%m-%d')
        sub_end_date = DateFormat(sub_date + relativedelta(months=1)).format("Y-m-d")
        accumulated = Transaction.objects.filter(
            business = business, Bkdate__gte = sessionInfo['start_date'], Bkdate__lt = sub_end_date
        ).exclude(item__code = 0).aggregate(
            input = Coalesce(Sum('Bkinput'),0),
            output = Coalesce(Sum('Bkoutput'),0))

        ym_range.append({'ym': ym, 'total_input': total['input'], 'total_output': total['output'], 'transaction': transaction, 'accumulated_input': accumulated['input'], 'accumulated_output': accumulated['output']})

    return render(request,'accounting/print_transaction.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'ym_range': ym_range, 'total_input': total_input, 'total_output':total_output})

@login_required(login_url='/')
def print_general_ledger(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    year2 = int(request.POST.get('year2', year))
    month2 = int(request.POST.get('month2', month))

    sessionInfo = session_info(str(year), str(month), business.session_month)

    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year2)+'-'+str(month2)+'-01', '%Y-%m-%d') + relativedelta(months=1)

    item_list = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3
    ).exclude(paragraph__subsection__code=0)
    for item in item_list :
        # 거래내역 가져오기
        transaction = []
        tr_in_item = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date, item = item).order_by('Bkdate','Bkid','Bkdivision')

        # 누계구하기
        tr_cumulative_total = Transaction.objects.filter(business = business, Bkdate__gte = sessionInfo['start_date'], Bkdate__lt = end_date, item = item).order_by('Bkdate','Bkid','Bkdivision').aggregate(Bkinput=Coalesce(Sum('Bkinput'),0), Bkoutput=Coalesce(Sum('Bkoutput'),0))

        # 월계구하기
        tr_total = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date, item = item).order_by('Bkdate','Bkid','Bkdivision').aggregate(Bkinput=Coalesce(Sum('Bkinput'),0), Bkoutput=Coalesce(Sum('Bkoutput'),0))

        tr_paginator = Paginator(tr_in_item, 40)
        for tr_page in range(1, tr_paginator.num_pages+1):
            transaction.append(tr_paginator.page(tr_page))
        if item.paragraph.subsection.type == "수입":
            filter_type = "revenue"
        elif item.paragraph.subsection.type == "지출":
            filter_type = "expenditure"

        # 예산액 가져오기
        budget = Budget.objects.filter(business = business, year = sessionInfo['year'], item=item, type__icontains=filter_type).order_by('type').last()
        prev_tr_total = Transaction.objects.filter(business=business, Bkdate__gte=sessionInfo['start_date'],
                                                   Bkdate__lt=start_date, item=item).aggregate(Bkinput=Coalesce(Sum('Bkinput'), 0), Bkoutput=Coalesce(Sum('Bkoutput'), 0))
        item.transaction = transaction
        item.cumulative_input = tr_cumulative_total['Bkinput']
        item.cumulative_output = tr_cumulative_total['Bkoutput']
        item.total_input = tr_total['Bkinput']
        item.total_output = tr_total['Bkoutput']

        if budget != None:
            item.budget = budget.price - prev_tr_total['Bkinput'] - prev_tr_total['Bkoutput']
        else:
            item.budget = 0
        item.balance = item.budget - item.total_input - item.total_output

    return render(request,'accounting/print_general_ledger.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'item_list': item_list})

@login_required(login_url='/')
def print_voucher(request, voucher_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    year2 = int(request.POST.get('year2', year))
    month2 = int(request.POST.get('month2', month))

    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(year2)+'-'+str(month2)+'-01', '%Y-%m-%d') + relativedelta(months=1)

    if voucher_type == 'revenue':
        filter_type = "수입"
        transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkinput__gt=0)
    else:
        filter_type = "지출"
        transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkoutput__gt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    data_list = []
    for ymd in ymd_list:
        if voucher_type == 'revenue':
            item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, transaction__Bkinput__gt=0, paragraph__subsection__type=filter_type).exclude(paragraph__subsection__code=0).distinct()
        else:
            item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, transaction__Bkoutput__gt=0, paragraph__subsection__type=filter_type).exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = []
            if voucher_type == 'revenue':
                tr_in_item = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkinput__gt=0)
            else:
                tr_in_item = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkoutput__gt=0)
            tr_sum = tr_in_item.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            tr_paginator = Paginator(tr_in_item, 20)
            for tr_page in range(1, tr_paginator.num_pages+1):
                transaction.append(tr_paginator.page(tr_page))

            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(item.sum)
            data_list.append({'date': ymd, 'item': item})

    return render(request,'accounting/print_voucher.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'year': year, 'month': month, 'year2': year2, 'month2': month2, 'data_list': data_list, 'voucher_type': voucher_type})

@login_required(login_url='/')
def print_voucher2(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    pk = request.GET.get('pk')
    transaction = get_object_or_404(Transaction, id=pk, business=business)
    if transaction.item.paragraph.subsection.type == "수입":
        transaction.sum_ko = readNumber(transaction.Bkinput)
    else:
        transaction.sum_ko = readNumber(transaction.Bkoutput)

    return render(request,'accounting/print_voucher2.html', {'settlement_management': 'active', 'master_login': request.session['master_login'], 'business': business, 'transaction': transaction})

@login_required(login_url='/')
def popup_returned_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    month = request.GET.get('month')
    Bkid = request.GET.get('Bkid')
    acctid = request.GET.get('acctid')
    Bkdate = request.GET.get('Bkdate')
    sessionInfo = session_info(Bkdate[:4], Bkdate[5:7], business.session_month)
    acct = get_object_or_404(Account, business=business, id=acctid)
    tr = TBLBANK.objects.get(Mid=business.owner.profile.user.username, Bkid=Bkid)
    if tr.Bkinput > 0:
        opposite_type = "지출"
    else :
        opposite_type = "수입"
    item_list = Item.objects.filter(
        paragraph__subsection__year=sessionInfo['year'],
        paragraph__subsection__type=opposite_type,
        paragraph__subsection__institution=business.type3
    ).exclude(paragraph__subsection__code=0)

    return render(request,'accounting/popup_returned_transaction.html', {'acct': acct, 'transaction': tr, 'item_list': item_list, 'month': month, 'Bkdate': Bkdate})

@login_required(login_url='/')
def regist_returned_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()

    if request.method == "POST":
        Bkid = request.POST.get('Bkid')
        tblbank_tr = TBLBANK.objects.get(Bkid=Bkid)
        Mid = request.user.username
        inoutType = request.POST.get('inoutType')
        itemId_list = request.POST.getlist('itemId_list')
        # print(len(itemId_list))
        subdivisionId_list = request.POST.getlist('subdivisionId_list')
        # print(len(subdivisionId_list))
        month = request.POST.get('month')
        Bkjukyo_list = request.POST.getlist('Bkjukyo_list')
        # print(len(Bkjukyo_list))
        Bkinout_list = request.POST.getlist('Bkinout_list')
        # print(len(Bkinout_list))
        Bkdate = request.POST.get('Bkdate')
        row = len(itemId_list) - 1

        try:
            close = Deadline.objects.get(business=business,year=Bkdate[:4],month=Bkdate[5:7])
            if close.regdatetime:
                return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
        except:
            pass

        sessionInfo = session_info(Bkdate[:4], Bkdate[5:7], business.session_month)
        start_date = datetime.datetime.strptime(Bkdate[:8] + '01', "%Y-%m-%d")
        a_month_ago = start_date - relativedelta(months=1)
        a_month_later = start_date + relativedelta(months=1)

        #--해당날짜 전월이월금 유무확인--
        try :
            premonth_transfer = Transaction.objects.filter(business=business, Bkdate=start_date).get(Bkdivision=0)
        except IndexError:
            return HttpResponse("<script>alert('IndexError');history.back();</script>")
        except Transaction.DoesNotExist:
            #--전월이월금 없는 경우 주계좌의 이전달 마지막 내역을 전월이월금으로 등록
            main = Account.objects.filter(business=business).get(main=True)
            data = TBLBANK.objects.filter(Bkacctno=main.account_number).filter(Bkdate__gte=a_month_ago).filter(Bkdate__lt=start_date).order_by('-Bkdate', '-Bkid').first()
            if data == None:
                return HttpResponse("<script>alert('주계좌의 전월잔액이 없습니다.');history.back();</script>")
            Transaction.objects.create(
                Bkid=data.Bkid,
                Bkdivision=0,
                Mid=Mid,
                business=business,
                Bkacctno=main.account_number,
                Bkname=main.bank.name,
                Bkdate=start_date,
                Bkjukyo="전월이월금",
                Bkinput=data.Bkjango,
                Bkoutput=0,
                Bkjango=data.Bkjango,
                item=Item.objects.get(
                    paragraph__subsection__year = sessionInfo['year'],
                    paragraph__subsection__institution=business.type3,
                    paragraph__subsection__code=0, paragraph__code=0, code=0),
                regdatetime=today
            )

        for r in range(row, 0, -1):
            latest_tr = Transaction.objects.filter(business=business, Bkdate__lte=Bkdate).order_by('-Bkdate', '-id').first()
            jango = latest_tr.Bkjango
            Bkinput = 0
            Bkoutput = 0
            Bkjango = 0
            if inoutType == "input":
                Bkinput = int(Bkinout_list[r]) * -1
                Bkjango = jango + Bkinput
            elif inoutType == "output":
                Bkoutput = int(Bkinout_list[r]) * -1
                Bkjango = jango - Bkoutput

            item = Item.objects.get(id=itemId_list[r])

            try:
                subdivision = Subdivision.objects.get(id=subdivisionId_list[r])
            except Subdivision.DoesNotExist:
                subdivision = None

            Transaction.objects.create(
                Bkid=Bkid,
                Bkdivision = r,
                Mid = Mid,
                Bkacctno = tblbank_tr.Bkacctno,
                Bkname = tblbank_tr.Bkname,
                Bkdate = Bkdate,
                Bkjukyo = "[분할]"+Bkjukyo_list[r],
                Bkinput = Bkinput,
                Bkoutput = Bkoutput,
                Bkjango = Bkjango,
                regdatetime = today,
                item = item,
                subdivision = subdivision,
                business = business,
            )

            update_list = Transaction.objects.filter(business=business, Bkdate__gt=Bkdate, Bkdate__lt=a_month_later)
            for update in update_list:
                if inoutType == "input":
                    update.Bkjango = int(update.Bkjango) + int(Bkinput)
                if inoutType == "output":
                    update.Bkjango = int(update.Bkjango) - int(Bkoutput)
                update.save()

            tr, created = TBLBANK.objects.get_or_create(business=business, Bkid=Bkid, Bkdivision=r, Mid=Mid, Bkacctno=tblbank_tr.Bkacctno, Bkname=tblbank_tr.Bkname, Bkdate=tblbank_tr.Bkdate)
            tr.sub_Bkjukyo="[분할]"+Bkjukyo_list[r]
            tr.regdatetime=today
            tr.item = item
            tr.subdivision = subdivision
            tr.save()

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def popup_transaction_direct(request):
    today = datetime.datetime.now()
    year = request.GET.get('year')
    month = request.GET.get('month')
    Bkdate = datetime.datetime.strptime(year+'-'+month+'-01', "%Y-%m-%d")
    business = get_object_or_404(Business, pk=request.session['business'])
    tblbankform = TblbankDirectForm(initial={'Bkdate':Bkdate})
    tblbankform.fields['item'].queryset = Item.objects.filter(
        paragraph__subsection__year = year,
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type="수입")
    return render(request,'accounting/popup_transaction_direct.html', {'transactionform': tblbankform, 'year': year, 'month': month})

@login_required(login_url='/')
def change_item_option(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    inoutType = request.GET.get('inoutType')
    year = request.GET.get('year')
    month = request.GET.get('month')
    Bkdate = datetime.datetime.strptime(year+'-'+month+'-01', "%Y-%m-%d")
    sessionInfo = session_info(year, month, business.session_month)
    html = '<option value="" selected="">---------</option>'
    if inoutType == "input":
        type_filter = "수입"
    elif inoutType == "output":
        type_filter = "지출"
    try :
        carryover_tr = Transaction.objects.filter(business=business, Bkdate=Bkdate).get(Bkdivision=0)
        items = Item.objects.filter(
            paragraph__subsection__year = sessionInfo['year'],
            paragraph__subsection__institution = business.type3,
            paragraph__subsection__type=type_filter
        ).exclude(paragraph__subsection__code=0)
    except Transaction.DoesNotExist:
        items = Item.objects.filter(
                paragraph__subsection__year = sessionInfo['year'],
                paragraph__subsection__institution = business.type3,
                paragraph__subsection__type=type_filter)
    for item in items:
        html += '<option value="'+str(item.id)+'">'+str(item)+'</option>'
    context = {'item': html}
    return HttpResponse(json.dumps(context), content_type="application/json")

@login_required(login_url='/')
def regist_transaction_direct(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    inout_type = request.POST.get('inout')
    remark = request.POST.get('remark')
    tblbankform = TblbankDirectForm(request.POST)
    Bkdate = request.POST.get('Bkdate')

    try:
        close = Deadline.objects.get(business=business, year=Bkdate[:4], month=Bkdate[5:7])
        if close.regdatetime:
            return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
    except:
        pass

    Bkid = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    if tblbankform.is_valid():
        tr = tblbankform.save(commit=False)
        tr.business = business
        tr.Bkid = Bkid
        if tr.item.context == "전월이월금":
            Bkdivision = 0
        else:
            Bkdivision = 1
        tr.Bkdivision = Bkdivision
        tr.Mid = business.owner.profile.user.username
        tr.regdatetime = today
        tr.direct = True
        tr.save()

        Bkdate = tr.Bkdate
        sessionInfo = session_info(str(tr.Bkdate.year), str(tr.Bkdate.month), business.session_month)
        start_date = datetime.datetime.strptime(DateFormat(tr.Bkdate).format("Y-m-01"), '%Y-%m-%d')
        # print(start_date) #2018-09-01 00:00:00
        # start_date2 = tr.Bkdate
        # print(start_date2) #2019-02-13 00:00:00+09:00
        a_month_ago = start_date - relativedelta(months=1)
        a_month_later = start_date + relativedelta(months=1)


        # --전월 이월금 등록--
        try:
            carryover_tr = Transaction.objects.filter(business=business, Bkdate=start_date).get(Bkdivision=0)
        # --전월이월금 없는 경우
        except Transaction.DoesNotExist:
            # --등록되는 거래가 전월이월금이면 transaction테이블에 등록 후 창닫기
            if tr.item.context == "전월이월금":
                Transaction.objects.create(
                    Bkid=Bkid, Bkdivision=0,
                    Mid=business.owner.profile.user.username, business=business,
                    Bkdate=start_date, Bkjukyo="전월이월금",
                    Bkinput=tr.Bkinput, Bkoutput=0,
                    Bkjango=tr.Bkinput, regdatetime=today,
                    item=Item.objects.get(
                        paragraph__subsection__year=sessionInfo['year'],
                        paragraph__subsection__institution=business.type3,
                        paragraph__subsection__code=0, paragraph__code=0, code=0),
                    remark=remark
                )
                return HttpResponse(
                    '<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')
            # --주계좌의 이전달 마지막 내역을 전월이월금으로 등록
            else:
                main_acct = Account.objects.filter(business=business, main=True).last()
                if main_acct is None:
                    account_number = 0
                    Bkname = ""
                else :
                    account_number = main_acct.account_number
                    Bkname = main_acct.bank.name
                last_tr = TBLBANK.objects.filter(Bkacctno=account_number, Bkdivision=1,
                                                 Bkdate__gte=a_month_ago, Bkdate__lt=start_date).order_by('Bkdate',
                                                                                                          'Bkid').last()
                # --전월 주계좌 거래내역 없으면 메시지 출력
                if last_tr == None:
                    tr.delete()
                    return HttpResponse("<script>alert('주계좌의 전월 거래내역이 없습니다. 전월이월금을 등록해주세요.');history.back();</script>")
                Transaction.objects.create(
                    Bkid=last_tr.Bkid, Bkdivision=0,
                    Mid=business.owner.profile.user.username, business=business,
                    Bkacctno=account_number,
                    Bkname=Bkname, Bkdate=start_date,
                    Bkjukyo="전월이월금", Bkinput=last_tr.Bkjango,
                    Bkoutput=0, Bkjango=last_tr.Bkjango,
                    item=Item.objects.get(
                        paragraph__subsection__year=sessionInfo['year'],
                        paragraph__subsection__institution=business.type3,
                        paragraph__subsection__code=0, paragraph__code=0, code=0),
                    regdatetime=today
                )
        # --전월이월금 등록 완료

        # --거래등록
        if tr.item.context == "전월이월금":
            return HttpResponse("<script>alert('해당월의 전년도 이월금은 이미 등록되었습니다.');history.back();</script>")

        latest_tr = Transaction.objects.filter(business=business, Bkdate__lte=Bkdate).order_by('-Bkdate', '-id').first()
        if inout_type == "input":
            Bkjango = latest_tr.Bkjango + tr.Bkinput
        elif inout_type == "output":
            Bkjango = latest_tr.Bkjango - tr.Bkoutput

        transaction = Transaction(
            Bkid=Bkid, Bkdivision=Bkdivision,
            Mid=business.owner.profile.user.username, business=business,
            Bkdate=Bkdate, Bkjukyo=tr.Bkjukyo,
            Bkinput=tr.Bkinput, Bkoutput=tr.Bkoutput,
            Bkjango=Bkjango, regdatetime=today,
            item=tr.item, remark=remark
        )
        transaction.save()

        update_list = Transaction.objects.filter(business=business, Bkdate__gt=Bkdate, Bkdate__lt=a_month_later)
        for update in update_list:
            if transaction.Bkinput > 0:
                update.Bkjango = update.Bkjango + transaction.Bkinput
            elif transaction.Bkoutput > 0:
                update.Bkjango = update.Bkjango - transaction.Bkoutput
            update.save()

    return HttpResponse(
        '<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def popup_transaction_edit(request):
    pk = request.GET.get('pk')
    business = get_object_or_404(Business, pk=request.session['business'])
    transaction = get_object_or_404(Transaction, id=pk, business=business)

    sessionInfo = session_info(str(transaction.Bkdate.year), str(transaction.Bkdate.month), business.session_month)

    tblbankform = TransactionEditForm(instance=transaction)
    tblbankform.fields['item'].queryset = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type=transaction.item.paragraph.subsection.type).exclude(code=0)
    inoutType = transaction.item.paragraph.subsection.type
    if inoutType == "수입":
        tblbankform.fields['Bkoutput'].widget.attrs['style'] = 'display:none'
        tblbankform.fields['Bkoutput'].widget.attrs['value'] = 0
    else:
        tblbankform.fields['Bkinput'].widget.attrs['style'] = 'display:none'
        tblbankform.fields['Bkinput'].widget.attrs['value'] = 0
    return render(request,'accounting/popup_transaction_edit.html', {'transactionform': tblbankform, 'year': transaction.Bkdate.year, 'month': transaction.Bkdate.month, 'pk': pk})

@login_required(login_url='/')
def edit_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    pk = request.POST.get('pk', '')
    year = request.POST.get('year', '')
    month = request.POST.get('month', '')
    transaction = get_object_or_404(Transaction, pk=pk)
    transactionform = TransactionEditForm(request.POST, instance=transaction)

    try:
        close = Deadline.objects.get(business=business,year=year,month=month)
        if close.regdatetime:
            return HttpResponse("<script>alert('해당월은 마감완료되었습니다.');history.back();</script>")
    except:
        pass

    #금액수정불가, 관항목만 수정가능
    if transactionform.is_valid():
        tr = transactionform.save(commit=False)
        if tr.Bkinput == None:
            tr.Bkinput = 0
        if tr.Bkoutput == None:
            tr.Bkoutput = 0
        tr.regdatetime = today
        tr.save()

        #TBLBANK테이블 동일한 항목 update
        update_tr = TBLBANK.objects.get(business=business, Bkid=tr.Bkid, Bkdivision=tr.Bkdivision)
        update_tr.item = tr.item
        update_tr.sub_Bkjukyo = tr.Bkjukyo
        update_tr.save()

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def print_returned_voucher(request, voucher_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))

    sessionInfo = session_info(str(year), str(month), business.session_month)

    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = start_date + relativedelta(months=1)

    if voucher_type == 'revenue':
        filter_type = "수입"
        transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkinput__lt=0)
    else:
        filter_type = "지출"
        transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkoutput__lt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()


    data_list = []
    for ymd in ymd_list:
        if voucher_type == 'revenue':
            item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, transaction__Bkinput__lt=0, paragraph__subsection__type=filter_type).exclude(paragraph__subsection__code=0).distinct()
        else:
            item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, transaction__Bkoutput__lt=0, paragraph__subsection__type=filter_type).exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            if voucher_type == 'revenue':
                transaction = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkinput__lt=0)
            else:
                transaction = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkoutput__lt=0)
            tr_sum = transaction.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(item.sum)
            data_list.append({'date': ymd, 'item': item})

    return render(request,'accounting/print_returned_voucher.html', {
        'settlement_management': 'active', 'master_login': request.session['master_login'],
        'business': business, 'year': year, 'month': month,
        'data_list': data_list, 'voucher_type': voucher_type
    })

@login_required(login_url='/')
def monthly_print(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    this_year = int(DateFormat(today).format("Y"))

    print_date = today - relativedelta(months = 1)
    year = int(DateFormat(print_date).format("Y"))
    month = int(DateFormat(print_date).format("m"))

    return render(request,'accounting/monthly_print.html', {'print_menu': 'active', 'monthly_print': 'active', 'master_login': request.session['master_login'], 'business': business, 'year_range': range(this_year, 1999, -1), 'month_range': range(1, 13), 'year': year, 'month': month})

@login_required(login_url='/')
def monthly_print_all(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    #--------기간설정----------
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))

    sessionInfo = session_info(str(year), str(month), business.session_month)

    start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
    end_date = start_date + relativedelta(months=1)
    print(start_date, end_date)

    #--------세입결산----------
    revenue_total_budget = 0
    revenue_total_sum = 0
    revenue_total_difference = 0

    filter_budget_type = getLatestBudgetType(business, year, "revenue")

    revenue_subsection_list = Subsection.objects.filter(
        year=sessionInfo['year'], institution=business.type3, type="수입"
    ).annotate(count=Count('paragraph__item')).exclude(count=0).exclude(code=0)
    revenue_paragraph_list = Paragraph.objects.filter(
        subsection__year=sessionInfo['year'], subsection__institution=business.type3, subsection__type="수입"
    ).annotate(count=Count('item')).exclude(count=0)
    revenue_item_list = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type = "수입"
    ).annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = sessionInfo['year'], then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0)).exclude(code=0)
    revenue_item_list2 = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type = "수입"
    ).annotate(
        total_sum=Coalesce(Sum(Case(
            When(transaction__business = business, then=Case(
            When(transaction__Bkdate__gte = start_date, then=Case(
            When(transaction__Bkdate__lt = end_date, then='transaction__Bkinput'))))))), 0)).exclude(code=0)

    for idx, val in enumerate(revenue_item_list):
        revenue_item_list[idx].total_sum = revenue_item_list2[idx].total_sum
        revenue_item_list[idx].total_difference = revenue_item_list[idx].total_budget-revenue_item_list[idx].total_sum
        revenue_total_budget += revenue_item_list[idx].total_budget
        revenue_total_sum += revenue_item_list[idx].total_sum
        revenue_total_difference += revenue_item_list[idx].total_difference

    #--------세출결산----------
    expenditure_total_budget = 0
    expenditure_total_sum = 0
    expenditure_total_difference = 0

    filter_budget_type = getLatestBudgetType(business, year, "expenditure")

    expenditure_subsection_list = Subsection.objects.filter(
        year=sessionInfo['year'], institution=business.type3, type="지출"
    ).annotate(count=Count('paragraph__item')).exclude(count=0)
    expenditure_paragraph_list = Paragraph.objects.filter(
        subsection__year = sessionInfo['year'], subsection__institution=business.type3, subsection__type="지출"
    ).annotate(count=Count('item')).exclude(count=0)
    expenditure_item_list = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type = "지출"
    ).annotate(
        total_budget=Coalesce(Sum(Case(
            When(budget__business = business, then=Case(When(budget__year = sessionInfo['year'], then=Case(When(budget__type = filter_budget_type, then='budget__price'))))))), 0))
    expenditure_item_list2 = Item.objects.filter(
        paragraph__subsection__year = sessionInfo['year'],
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__type = "지출"
    ).annotate(
        total_sum=Coalesce(Sum(Case(
            When(transaction__business = business, then=Case(
            When(transaction__Bkdate__gte = start_date, then=Case(
            When(transaction__Bkdate__lt = end_date, then='transaction__Bkoutput'))))))), 0))

    for idx, val in enumerate(expenditure_item_list):
        expenditure_item_list[idx].total_sum = expenditure_item_list2[idx].total_sum
        expenditure_item_list[idx].total_difference = expenditure_item_list[idx].total_budget-expenditure_item_list[idx].total_sum
        expenditure_total_budget += expenditure_item_list[idx].total_budget
        expenditure_total_sum += expenditure_item_list[idx].total_sum
        expenditure_total_difference += expenditure_item_list[idx].total_difference

    #--------현금출납부----------

    transaction_list = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date).order_by('Bkdate','Bkid','Bkdivision').exclude(item=None)

    ym_list = []
    for transaction in transaction_list:
        ym_list.append(str(transaction.Bkdate)[:7])
    ym_list = list(set(ym_list))
    ym_list.sort()

    ym_range = []
    total_input = 0
    total_output = 0
    for ym in ym_list:
        transaction = []
        ym_tr = Transaction.objects.filter(business = business, Bkdate__year=ym[:4], Bkdate__month=ym[-2:]).order_by('Bkdate','Bkid')
        tr_paginator = Paginator(ym_tr, 40)
        for tr_page in range(1, tr_paginator.num_pages+1):
            transaction.append(tr_paginator.page(tr_page))

        total = Transaction.objects.filter(business=business, Bkdate__year=ym[:4], Bkdate__month=ym[-2:]).aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
        sub_date =  datetime.datetime.strptime(ym[:4]+'-'+ym[-2:]+'-01', '%Y-%m-%d')
        sub_end_date = DateFormat(sub_date + relativedelta(months=1)).format("Y-m-d")
        accumulated = Transaction.objects.filter(
            business = business, Bkdate__gte = sessionInfo['start_date'], Bkdate__lt = sub_end_date
        ).exclude(item__code = 0).aggregate(
            input = Coalesce(Sum('Bkinput'), 0),
            output = Coalesce(Sum('Bkoutput'), 0))

        ym_range.append({'ym': ym, 'total_input': total['input'], 'total_output': total['output'], 'transaction': transaction, 'accumulated_input': accumulated['input'], 'accumulated_output': accumulated['output']})

    #--------총계정원장----------
    general_ledger_list = Item.objects.filter(
        paragraph__subsection__year=sessionInfo['year'],
        paragraph__subsection__institution = business.type3
    ).exclude(paragraph__subsection__code=0)
    for general_ledger in general_ledger_list :
        # 거래내역 가져오기
        transaction = []
        tr_in_general_ledger = Transaction.objects.filter(business = business, Bkdate__gte = start_date, Bkdate__lt = end_date, item = general_ledger).order_by('Bkdate','Bkid','Bkdivision')

        # 누계구하기
        tr_cumulative_total = Transaction.objects.filter(
            business = business, Bkdate__gte = sessionInfo['start_date'], Bkdate__lt = end_date, item = general_ledger
        ).order_by('Bkdate', 'Bkid', 'Bkdivision').aggregate(Bkinput = Coalesce(Sum('Bkinput'), 0), Bkoutput=Coalesce(Sum('Bkoutput'), 0))

        # 월계구하기
        tr_total = Transaction.objects.filter(
            business = business, Bkdate__gte = start_date, Bkdate__lt = end_date, item = general_ledger
        ).order_by('Bkdate', 'Bkid', 'Bkdivision').aggregate(Bkinput = Coalesce(Sum('Bkinput'),0), Bkoutput=Coalesce(Sum('Bkoutput'), 0))

        tr_paginator = Paginator(tr_in_general_ledger, 40)
        for tr_page in range(1, tr_paginator.num_pages+1):
            transaction.append(tr_paginator.page(tr_page))
        if general_ledger.paragraph.subsection.type == "수입":
            filter_type = "revenue"
        elif general_ledger.paragraph.subsection.type == "지출":
            filter_type = "expenditure"

        # 예산액 가져오기
        budget = Budget.objects.filter(business = business, year = sessionInfo['year'], item=general_ledger, type__icontains=filter_type).order_by('type').last()
        general_ledger.transaction = transaction
        general_ledger.cumulative_input = tr_cumulative_total['Bkinput']
        general_ledger.cumulative_output = tr_cumulative_total['Bkoutput']
        general_ledger.total_input = tr_total['Bkinput']
        general_ledger.total_output = tr_total['Bkoutput']
        if budget != None:
            general_ledger.budget = budget.price
        else:
            general_ledger.budget = 0
        general_ledger.balance = general_ledger.budget - general_ledger.cumulative_input - general_ledger.cumulative_output

    #--------수입결의서----------
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkinput__gt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    revenue_voucher_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, paragraph__subsection__type="수입").exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = []
            tr_in_item = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkinput__gt=0)
            tr_paginator = Paginator(tr_in_item, 20)
            for tr_page in range(1, tr_paginator.num_pages+1):
                transaction.append(tr_paginator.page(tr_page))
            tr_sum = tr_in_item.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(item.sum)
            revenue_voucher_list.append({'date': ymd, 'item': item})

    #--------지출결의서----------
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkoutput__gt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    expenditure_voucher_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, paragraph__subsection__type="지출").exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = []
            tr_in_item = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkoutput__gt=0)
            tr_paginator = Paginator(tr_in_item, 20)
            for tr_page in range(1, tr_paginator.num_pages+1):
                transaction.append(tr_paginator.page(tr_page))
            tr_sum = tr_in_item.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(item.sum)
            expenditure_voucher_list.append({'date': ymd, 'item': item})

    #--------수입반납결의서----------
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkinput__lt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    revenue_returned_voucher_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, transaction__Bkinput__lt=0, paragraph__subsection__type="수입").exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkinput__lt=0)
            tr_sum = transaction.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(item.sum)
            revenue_returned_voucher_list.append({'date': ymd, 'item': item})

    #--------지출반납결의서----------
    transaction_list = Transaction.objects.filter(business=business, Bkdate__gte=start_date, Bkdate__lt=end_date, Bkoutput__lt=0)

    ymd_list = []
    for transaction in transaction_list:
        ymd_list.append(transaction.Bkdate)
    ymd_list = list(set(ymd_list))
    ymd_list.sort()

    expenditure_returned_voucher_list = []
    for ymd in ymd_list:
        item_list = Item.objects.filter(transaction__business=business, transaction__Bkdate=ymd, transaction__Bkoutput__lt=0, paragraph__subsection__type="지출").exclude(paragraph__subsection__code=0).distinct()
        for item in item_list:
            transaction = Transaction.objects.filter(business=business, Bkdate=ymd, item=item, Bkoutput__lt=0)
            tr_sum = transaction.aggregate(input=Coalesce(Sum('Bkinput'),0), output=Coalesce(Sum('Bkoutput'),0))
            item.transaction = transaction
            item.sum = tr_sum['input'] + tr_sum['output']
            item.sum_ko = readNumber(item.sum)
            expenditure_returned_voucher_list.append({'date': ymd, 'item': item})

    return render(request,'accounting/monthly_print_all.html', {
        'settlement_management': 'active', 'master_login': request.session['master_login'],
        'business': business, 'year': year, 'month': month,
        'revenue_subsection_list': revenue_subsection_list, 'revenue_paragraph_list': revenue_paragraph_list, 'revenue_item_list': revenue_item_list, 'revenue_total_budget': revenue_total_budget, 'revenue_total_sum': revenue_total_sum, 'revenue_total_difference': revenue_total_difference,
        'expenditure_subsection_list': expenditure_subsection_list, 'expenditure_paragraph_list': expenditure_paragraph_list, 'expenditure_item_list': expenditure_item_list, 'expenditure_total_budget': expenditure_total_budget, 'expenditure_total_sum': expenditure_total_sum, 'expenditure_total_difference': expenditure_total_difference,
        'ym_range': ym_range, 'total_input': total_input, 'total_output':total_output,
        'general_ledger_list': general_ledger_list,
        'revenue_voucher_list': revenue_voucher_list,
        'expenditure_voucher_list': expenditure_voucher_list,
        'revenue_returned_voucher_list': revenue_returned_voucher_list,
        'expenditure_returned_voucher_list': expenditure_returned_voucher_list
    })

@login_required(login_url='/')
def print_trial_balance(request):
    if request.method == "POST":
        business = get_object_or_404(Business, pk=request.session['business'])
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        if request.POST.get('year2'):
            year2 = int(request.POST.get('year2'))
            month2 = int(request.POST.get('month2'))
        else:
            year2 = year
            month2 = month

        sessionInfo = session_info(str(year), str(month), business.session_month)

        start_date = datetime.datetime.strptime(str(year)+'-'+str(month)+'-01', '%Y-%m-%d')
        end_date = datetime.datetime.strptime(str(year2)+'-'+str(month2)+'-01', '%Y-%m-%d') + relativedelta(months=1)

        item_list = Item.objects.filter(
            paragraph__subsection__year = sessionInfo['year'],
            paragraph__subsection__institution = business.type3
        ).order_by('paragraph__subsection__type', 'paragraph__subsection__code', 'paragraph__code', 'code').exclude(code=0)

        item_list.cumulative_income_total = 0       # 수입누계 총합
        item_list.income_total = 0                  # 수입금액 총합
        item_list.expenditure_total = 0             # 지출금액 총합
        item_list.cumulative_expenditure_total = 0  # 지출누계 총합
        for idx, item in enumerate(item_list):
            if item.paragraph.subsection.type == '수입':
                # 목별 수입누계
                item_list[idx].cumulative_income = getTransactionSumByItem(
                            business, item, sessionInfo['start_date'], end_date)
                item_list.cumulative_income_total += item_list[idx].cumulative_income

                # 목별 수입금액
                item_list[idx].income = getTransactionSumByItem(
                                            business, item, start_date, end_date)
                item_list.income_total += item_list[idx].income
            elif item.paragraph.subsection.type == '지출':
                # 목별 지출금액
                item_list[idx].expenditure = getTransactionSumByItem(
                                                business, item, start_date, end_date)
                item_list.expenditure_total += item_list[idx].expenditure

                # 목별 수입누계
                item_list[idx].cumulative_expenditure = getTransactionSumByItem(
                            business, item, sessionInfo['start_date'], end_date)
                item_list.cumulative_expenditure_total += item_list[idx].cumulative_expenditure

        itemList = []
        item_paginator = Paginator(item_list, 40)
        for item_page in range(1, item_paginator.num_pages+1):
            itemList.append(item_paginator.page(item_page))

    return render(request, 'accounting/print_trial_balance.html', {
        'settlement_management': 'active', 'trial_balance_page': 'active',
        'year': year, 'month': month, 'year2': year2, 'month2': month2,
        'item_list': item_list, 'itemList' : itemList
    })

@login_required(login_url='/')
def close_list(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "GET":
        today = datetime.datetime.now()
        this_year = today.year
        selected_year = int(request.GET.get('year', this_year))

        sessionInfo = session_info(str(selected_year), business.session_month, business.session_month)

        ym_list = []
        for m in range(0, 12):
            ym = sessionInfo['start_date'] + relativedelta(months=m)
            try:
                get_deadline = Deadline.objects.get(business=business, year=ym.year, month=ym.month)
                if get_deadline.regdatetime:
                    close = 1
                else:
                    close = 0
            except Deadline.DoesNotExist:
                    close = 0
            ym_list.append({'Bkdate__year': ym.year, 'Bkdate__month': ym.month, 'close': close})

    return render(request,'accounting/close_list.html', {'ym_list': ym_list, 'year_range': range(this_year, 1999, -1), 'selected_year': selected_year, 'accounting_management': 'active','close_list': 'active'})

@login_required(login_url='/')
def regist_close(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    if request.method == "POST":
        ym = request.POST.get('ym')
        year = ym[:4]
        month = ym[5:]

        main_acct = Account.objects.get(business=business, main=True)
        try:
            tblbank_last_jango = TBLBANK.objects.filter(Bkacctno=main_acct.account_number, Bkdivision=1, Bkdate__year=year, Bkdate__month=month).order_by('Bkdate','id').last().Bkjango
            transaction_last_jango = Transaction.objects.filter(business=business, Bkdivision=1, Bkdate__year=year, Bkdate__month=month).order_by('Bkdate', 'id').last().Bkjango
        except TBLBANK.DoesNotExist:
            return HttpResponse("<script>alert('등록된 거래가 없습니다.');history.back();</script>")
        except AttributeError:
            return HttpResponse("<script>alert('등록된 거래가 없습니다.');history.back();</script>")

        if tblbank_last_jango == transaction_last_jango:
            deadline, created = Deadline.objects.get_or_create(business=business, year=year, month=month)
            deadline.regdatetime = today
            deadline.save()
        else:
            return HttpResponse("<script>alert('주계좌의 잔액과 거래내역의 잔액이 일치하지 않습니다.');history.back();</script>")

    response = redirect('close_list')
    if business.type3_id == "어린이집" and int(month) < 3:
        response['Location'] += '?year='+str(int(year)-1) #회기년도
    else:
        response['Location'] += '?year='+year   #회기년도
    return response

@login_required(login_url='/')
def undo_close(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    if request.method == "POST":
        ym = request.POST.get('ym')
        year = ym[:4]
        month = ym[5:]

        try:
            deadline = Deadline.objects.get(business=business,year=year, month=month)
            deadline.regdatetime = None
            deadline.save()
        except:
            pass

    response = redirect('close_list')
    if business.type3_id == "어린이집" and int(month) < 3:
        response['Location'] += '?year='+str(int(year)-1) #회기년도
    else:
        response['Location'] += '?year='+year   #회기년도
    return response

@login_required(login_url='/')
def authkey_list(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    return render(request, 'accounting/authkey_list.html', {'business_management': 'active', 'authkey_page': 'active', 'business': business})

@login_required(login_url='/')
def authkey_edit(request):
    from .forms import AuthKeyForm
    business = get_object_or_404(Business, pk=request.session['business'])
    if request.method == "POST":
        form = AuthKeyForm(request.POST, instance=business)
        if form.is_valid():
            form.save()
            return redirect('authkey_list')
    else:
        form = AuthKeyForm()
    return render(request, 'accounting/authkey_edit.html', {'business_management': 'active', 'authkey_page': 'active', 'business': business, 'form': form})

@login_required(login_url='/')
def print_budget_all(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))

    general = _budget_general(request, year)

    revenue_budget = _budget_content(request, year, 'revenue')

    expenditure_budget = _budget_content(request, year, 'expenditure')

    return render(request, 'accounting/print_budget_all.html', {
        'business': business, 'year': year,
        'general': general, 'revenue_budget': revenue_budget,
        'expenditure_budget': expenditure_budget
    })

def _budget_general(request, year):
    business = get_object_or_404(Business, pk=request.session['business'])

    is_total = 0    #세입-관
    is_xtotal = 0   #세입-전년도
    is_dtotal = 0   #세입-증감
    os_total = 0    #세출-관
    os_xtotal = 0   #세출-전년도
    os_dtotal = 0   #세출-증감

    isubsection_list = subsection_info(year, business.type3, 'i')
    for subsection in isubsection_list:
        subsection.s_total = Budget.objects.filter(business=business, year=year, item__paragraph__subsection = subsection, type="revenue").aggregate(total=Coalesce(Sum('price'), 0))['total']
        subsection.xs_total = Budget.objects.filter(business=business, year=year-1, item__paragraph__subsection__code = subsection.code, type="revenue").aggregate(total=Coalesce(Sum('price'), 0))['total']
        subsection.ds_total = subsection.s_total - subsection.xs_total
        is_total += subsection.s_total
        is_xtotal += subsection.xs_total
        is_dtotal += subsection.ds_total

    osubsection_list = subsection_info(year, business.type3, 'o')
    for subsection in osubsection_list:
        subsection.s_total = Budget.objects.filter(business=business, year=year, item__paragraph__subsection = subsection, type="expenditure").aggregate(total=Coalesce(Sum('price'), 0))['total']
        subsection.xs_total = Budget.objects.filter(business=business, year=year-1, item__paragraph__subsection__code = subsection.code, type="expenditure").aggregate(total=Coalesce(Sum('price'), 0))['total']
        subsection.ds_total = subsection.s_total - subsection.xs_total
        os_total += subsection.s_total
        os_xtotal += subsection.xs_total
        os_dtotal += subsection.ds_total

    general = {
        'isubsection_list': isubsection_list, 'osubsection_list': osubsection_list,
        'is_total': is_total, 'is_xtotal': is_xtotal, 'is_dtotal': is_dtotal,
        'os_total': os_total, 'os_xtotal': os_xtotal, 'os_dtotal': os_dtotal,
    }
    return general

def _budget_content(request, year, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])

    # 예산전체출력은 본예산 기준(추경예산X)
    if budget_type in ['revenue']:
        stype_filter = '수입'
    elif budget_type in ['expenditure']:
        stype_filter='지출'

    # 총합 : 예산액, 전년도 예산액, 비교증감
    total = Budget.objects.filter(business=business, year=year, item__paragraph__subsection__type=stype_filter, type=budget_type).aggregate(total=Coalesce(Sum('price'), 0))['total']
    xtotal = Budget.objects.filter(business=business, year=year-1, item__paragraph__subsection__type=stype_filter, type=budget_type).aggregate(total=Coalesce(Sum('price'), 0))['total']
    dtotal = total - xtotal

    subsection_list = Subsection.objects.filter(year=year, type=stype_filter, institution=business.type3).exclude(code=0)
    for subsection in subsection_list:
        subsection.s_total = Budget.objects.filter(business=business, year=year, item__paragraph__subsection = subsection, type=budget_type).aggregate(total=Coalesce(Sum('price'), 0))['total']
        subsection.xs_total = Budget.objects.filter(business=business, year=year-1, item__paragraph__subsection__code = subsection.code, type=budget_type).aggregate(total=Coalesce(Sum('price'), 0))['total']
        subsection.ds_total = subsection.s_total - subsection.xs_total
        # print("관(", subsection.code, ") : ", subsection.s_total, subsection.xs_total, subsection.ds_total)

        paragraph_list = Paragraph.objects.filter(subsection=subsection)
        for paragraph in paragraph_list:
            paragraph.p_total = Budget.objects.filter(business=business, year=year, item__paragraph=paragraph, type=budget_type).aggregate(total=Coalesce(Sum('price'), 0))['total']
            paragraph.xp_total = Budget.objects.filter(business=business, year=year-1, item__paragraph__subsection__code = subsection.code, item__paragraph__code = paragraph.code, type=budget_type).aggregate(total=Coalesce(Sum('price'), 0))['total']
            paragraph.dp_total = paragraph.p_total - paragraph.xp_total
            # print("관항(", subsection.code,paragraph.code, ") : ", paragraph.p_total, paragraph.xp_total, paragraph.dp_total)

            item_list = Item.objects.filter(paragraph=paragraph)
            for item in item_list:
                budget = Budget.objects.get(business=business, year=year, item=item, type=budget_type)
                item.i_total = budget.price
                try :
                    item.xi_total = Budget.objects.get(business=business, year=year-1, item__paragraph__subsection__code = subsection.code, item__paragraph__code = paragraph.code, item__code=item.code, type=budget_type).price
                except :
                    item.xi_total = 0
                item.di_total = item.i_total - item.xi_total
                # print("관항목(", subsection.code, paragraph.code, item.code, ") : ", item.i_total, item.xi_total, item.di_total)

                percent_list = []
                sub_columns = ['item','context','unit_price','cnt','months','percent','sub_price']
                context_list = budget.context.split("|")
                unit_price_list = budget.unit_price.split("|")
                cnt_list = budget.cnt.split("|")
                months_list = budget.months.split("|")
                if budget.percent is not None:
                    percent_list = budget.percent.split("|")
                sub_price_list = budget.sub_price.split("|")
                row_list = []
                for idx, val in enumerate(context_list):
                    r = []
                    if val != None:
                        r.append(budget.item.id)
                        r.append(context_list[idx])
                        r.append(unit_price_list[idx])
                        r.append(cnt_list[idx])
                        r.append(months_list[idx])
                        if budget.percent is not None:
                            r.append(percent_list[idx])
                        else:
                            r.append('')
                        r.append(sub_price_list[idx])
                    row_list.append(r)
                    item.sub_data = [ dict(zip(sub_columns,row)) for row in row_list ]
            paragraph.item_list = item_list
        subsection.paragraph_list = paragraph_list

    page_list = []
    sub_row = []
    for index, subsection in enumerate(subsection_list):
        sub_row.append(subsection)
        if (budget_type in ['revenue', 'supplementary_revenue', 'supplementary_revenue']) and (index+1 in [3,9]):
            page_list.append(sub_row)
            sub_row = []
        elif (budget_type in ['expenditure', 'supplementary_expenditure', 'supplementary_expenditure']) and (index+1 in [1,2,5,10]):
            page_list.append(sub_row)
            sub_row = []

    budget_content = {
        'budget_type': budget_type, 'page_list': page_list,
        'total': total, 'xtotal': xtotal, 'dtotal': dtotal
    }
    return budget_content

@login_required(login_url='/')
def print_settlement_all(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))

    general = _settlement_general(request, year)

    revenue_settlement = _settlement_content(request, year, "revenue")
    expenditure_settlement = _settlement_content(request, year, "expenditure")

    session = session_info(str(year), business.session_month, business.session_month)
    bojo_list = getTransactionListByName(business, "수입", session['start_date'], session['end_date'], "보조금", "N")

    return render(request, 'accounting/print_settlement_all.html', {
        'business': business, 'year': year, 'general': general, 'bojo_list': bojo_list,
        'revenue_settlement': revenue_settlement, 'expenditure_settlement': expenditure_settlement
    })

def _settlement_general(request, year):
    business = get_object_or_404(Business, pk=request.session['business'])
    session = session_info(str(year), business.session_month, business.session_month)
    prevSession = session_info(str(year - 1), business.session_month, business.session_month)

    revenue_budget_total = 0        # 세입예산 총합
    revenue_settlement_total = 0    # 세입결산 총합
    revenue_diff_total = 0          # 세입차액 총합
    expenditure_budget_total = 0    # 세입예산 총합
    expenditure_settlement_total = 0 # 세입결산 총합
    expenditure_diff_total = 0      # 세입차액 총합

    # 관목록(수입)
    isubsection_list = subsection_info(year, business.type3, 'i')
    revenue_budget_type = getLatestBudgetType(business, year, "revenue")
    for subsection in isubsection_list:
        subsection.budget_amount = getBudgetSumBySubsection(business, year, subsection, revenue_budget_type)
        subsection.settlement_amount = \
            getTransactionSumBySubsection(business, subsection, session['start_date'], session['end_date'])
        subsection.diff = subsection.budget_amount - subsection.settlement_amount
        revenue_budget_total += subsection.budget_amount
        revenue_settlement_total += subsection.settlement_amount
        revenue_diff_total += subsection.diff

    # 관목록(지출)
    osubsection_list = subsection_info(year, business.type3, 'o')
    expenditure_budget_type = getLatestBudgetType(business, year, "expenditure")
    for subsection in osubsection_list:
        subsection.budget_amount = getBudgetSumBySubsection(business, year, subsection, expenditure_budget_type)
        subsection.settlement_amount = \
            getTransactionSumBySubsection(business, subsection, session['start_date'], session['end_date'])
        subsection.diff = subsection.budget_amount - subsection.settlement_amount
        expenditure_budget_total += subsection.budget_amount
        expenditure_settlement_total += subsection.settlement_amount
        expenditure_diff_total += subsection.diff

    """ 내용 확인 후 적용
    # 적립금 코드(511)
    reserve_item = Item.objects.get(
        paragraph__subsection__institution = business.type3,
        paragraph__subsection__year = year, paragraph__subsection__type = "지출",
        paragraph__subsection__code = 5, paragraph__code = 1, code = 1
    )
    this_year_reserve = getTransactionSumByCode(business, reserve_item, session['start_date'], session['end_date'])
    prev_year_reserve = getTransactionSumByCode(business, reserve_item, prevSession['start_date'], prevSession['end_date'])

    # 적립금 처분수입(611)
    reserve_income_item = Item.objects.get(
        paragraph__subsection__institution=business.type3,
        paragraph__subsection__year=year, paragraph__subsection__type="수입",
        paragraph__subsection__code=6, paragraph__code=1, code=1
    )
    reserve_income = getTransactionSumByCode(business, reserve_income_item, session['start_date'], session['end_date'])
    """

    general = {
        'isubsection_list': isubsection_list, 'osubsection_list': osubsection_list,
        'revenue_budget_total': revenue_budget_total, 'expenditure_budget_total': expenditure_budget_total,
        'revenue_settlement_total': revenue_settlement_total, 'expenditure_settlement_total': expenditure_settlement_total,
        'revenue_diff_total': revenue_diff_total, 'expenditure_diff_total': expenditure_diff_total
        #'this_year_reserve': this_year_reserve, 'prev_year_reserve': prev_year_reserve,
        #'reserve_income': reserve_income
    }
    return general

def _settlement_content(request, year, budget_type):
    business = get_object_or_404(Business, pk=request.session['business'])
    session = session_info(str(year), business.session_month, business.session_month)

    total_budget = 0
    total_settlement = 0
    total_difference = 0

    if budget_type == "revenue":
        filter_type = "수입"
        budget_type = getLatestBudgetType(business, year, "revenue")
    elif budget_type == "expenditure":
        filter_type = "지출"
        budget_type = getLatestBudgetType(business, year, "expenditure")

    item_list = Item.objects.filter(
        paragraph__subsection__year=year,
        paragraph__subsection__institution=business.type3,
        paragraph__subsection__type=filter_type
    ).exclude(code=0)

    for idx, item in enumerate(item_list):
        item.budget_amount = getBudgetSumByItem(business, year, item, budget_type)
        total_budget += item.budget_amount
        item.settlement_amount = getTransactionSumByItem(business, item, session['start_date'], session['end_date'])
        total_settlement += item.settlement_amount
        item.difference_amount = item.budget_amount - item.settlement_amount
        total_difference += item.difference_amount

        # template에서 편하게 쓰기위해 추가
        item.scode = item.paragraph.subsection.code
        item.pcode = item.paragraph.code
        item.scontext = item.paragraph.subsection.context
        item.pcontext = item.paragraph.context

    content = {
        'item_list': item_list, 'total_budget': total_budget,
        'total_settlement': total_settlement, 'total_difference': total_difference
    }
    return content

#--------------파일다운로드-------------
from .models import UploadFile

@login_required(login_url='/')
def file_download(request):
    upload_files = UploadFile.objects.filter(user__is_staff=True)

    return render(request, 'accounting/file_download.html', {'download': 'active', 'upload_files': upload_files})

@login_required(login_url='/')
def popup_upload(request):
    type = request.GET.get('type','')
    return render(request, 'accounting/popup_upload.html', {'type': type})


from openpyxl import load_workbook

@login_required(login_url='/')
def upload_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    ymd = DateFormat(today).format("ymdHis")
    upfile = request.FILES.get('file')
    upload_type = request.POST.get('type')

    if upfile == None:
        return HttpResponse("<script>alert('파일을 선택해주세요.');history.back();</script>")
    upfile.name = str(ymd)+'_'+business.name+'_거래내역.xlsx'
    created_file = UploadFile.objects.create(title=str(ymd)+'_'+business.name+'_거래내역', file=upfile, user=request.user)

    wb = load_workbook(filename='./media/'+upfile.name)
    sheet = wb.worksheets[0]
    try:
        Bkacct = Account.objects.get(business=business, account_number=sheet['B1'].value)
    except:
        return HttpResponse("<script>alert('등록되지 않은 계좌입니다. 엑셀파일의 계좌번호를 확인해주세요.');history.back();</script>")

    num = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    for index in range(5, sheet.max_row + 1):
        if sheet['A'+str(index)] and sheet['B'+str(index)] and sheet['E'+str(index)] and (sheet['C'+str(index)] or sheet['D'+str(index)]):
            if sheet['C'+str(index)].value and not sheet['D'+str(index)].value:
                if sheet['C'+str(index)].value > 0:
                    Bkinput = sheet['C'+str(index)].value
                    Bkoutput = 0
                elif sheet['C'+str(index)].value < 0:
                    Bkinput = 0
                    Bkoutput = sheet['C'+str(index)].value * -1
            elif sheet['D'+str(index)].value and not sheet['C'+str(index)].value:
                if sheet['D'+str(index)].value > 0:
                    Bkinput = 0
                    Bkoutput = sheet['D'+str(index)].value
                elif sheet['D'+str(index)].value < 0:
                    Bkinput = sheet['D'+str(index)].value * -1
                    Bkoutput = 0

            datetimecell = sheet['A'+str(index)].value
            if type(datetimecell) == datetime.datetime:
                Bkdate = datetimecell
            else:
                try:
                    datetimecell = datetimecell.replace(".","-").replace("/","-")
                    Bkdate = datetime.datetime.strptime(datetimecell, '%Y-%m-%d %H:%M:%S')
                except:
                    created_file.delete()
                    return HttpResponse("<script>alert('"+str(index-5)+"개의 거래 등록 완료.<br>"+str(index)+"행의 날짜형식이 맞지 않습니다.');history.back();</script>")

            TBLBANK.objects.create(
                Bkid=num,
                Bkdivision=1,
                Mid=request.user.username,
                Bkacctno=Bkacct.account_number,
                Bkname=Bkacct.bank.name,
                Bkdate=Bkdate,
                Bkjukyo=sheet['B'+str(index)].value,
                Bkinput=Bkinput,
                Bkoutput=Bkoutput,
                Bkjango=sheet['E'+str(index)].value,
                business=business
            )
        num += 1

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def upload_transaction2(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    ymd = DateFormat(today).format("ymdHis")
    upfile = request.FILES.get('file')
    upload_type = request.POST.get('type')

    if upfile == None:
        return HttpResponse("<script>alert('파일을 선택해주세요.');history.back();</script>")
    upfile.name = str(ymd)+'_'+business.name+'_거래내역.xlsx'
    created_file = UploadFile.objects.create(title=str(ymd)+'_'+business.name+'_거래내역', file=upfile, user=request.user)

    wb = load_workbook(filename='./media/'+upfile.name)
    sheet = wb.worksheets[0]
    try:
        Bkacct = Account.objects.get(business=business, account_number=sheet['B1'].value)
    except:
        created_file.delete()
        return HttpResponse("<script>alert('등록되지 않은 계좌입니다. 엑셀파일의 계좌번호를 확인해주세요.');history.back();</script>")

    num = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    for index in range(5, sheet.max_row + 1):
        if sheet['A'+str(index)] and sheet['B'+str(index)] and sheet['C'+str(index)] and sheet['F'+str(index)] and (sheet['D'+str(index)] or sheet['E'+str(index)]):
            if sheet['D'+str(index)].value and not sheet['E'+str(index)].value:
                if sheet['D'+str(index)].value > 0:
                    Bkinput = sheet['D'+str(index)].value
                    Bkoutput = 0
                elif sheet['D'+str(index)].value < 0:
                    Bkinput = 0
                    Bkoutput = sheet['D'+str(index)].value * -1
            elif sheet['E'+str(index)].value and not sheet['D'+str(index)].value:
                if sheet['E'+str(index)].value > 0:
                    Bkinput = 0
                    Bkoutput = sheet['E'+str(index)].value
                elif sheet['E'+str(index)].value < 0:
                    Bkinput = sheet['E'+str(index)].value * -1
                    Bkoutput = 0

            datecell = sheet['A'+str(index)].value
            if type(datecell) == datetime.datetime:
                date = DateFormat(sheet['A'+str(index)].value).format("Y-m-d ")
            else:
                try:
                    datecell = datecell.replace(".","-").replace("/","-")
                    date = datecell+" "
                except:
                    created_file.delete()
                    return HttpResponse("<script>alert('"+str(index-5)+"개의 거래 등록 완료.<br>"+str(index)+"행의 날짜형식이 맞지 않습니다.');history.back();</script>")

            timecell = sheet['B'+str(index)].value
            if type(timecell) == datetime.datetime:
                time = DateFormat(sheet['B'+str(index)].value).format("H:i:s")
            else:
                try:
                    time = timecell
                except:
                    created_file.delete()
                    return HttpResponse("<script>alert('"+str(index-5)+"개의 거래 등록 완료.<br>"+str(index)+"행의 시간형식이 맞지 않습니다.');history.back();</script>")

            try:
                Bkdate = datetime.datetime.strptime(date+time, '%Y-%m-%d %H:%M:%S')
            except:
                created_file.delete()
                return HttpResponse("<script>alert('"+str(index-5)+"개의 거래 등록 완료.<br>"+str(index)+"행의 날짜 혹은 시간형식이 맞지 않습니다.');history.back();</script>")

            TBLBANK.objects.create(
                Bkid=num,
                Bkdivision=1,
                Mid=request.user.username,
                Bkacctno=Bkacct.account_number,
                Bkname=Bkacct.bank.name,
                Bkdate=Bkdate,
                Bkjukyo=sheet['C'+str(index)].value,
                Bkinput=Bkinput,
                Bkoutput=Bkoutput,
                Bkjango=sheet['F'+str(index)].value,
                business=business
            )
        num += 1

    return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')


from django.db import transaction

@login_required(login_url='/')
def upload_voucher(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    today = datetime.datetime.now()
    ymd = DateFormat(today).format("ymdHis")
    upfile = request.FILES.get('file')

    if upfile == None:
        return HttpResponse("<script>alert('파일을 선택해주세요.');history.back();</script>")
    upfile.name = str(ymd)+'_'+business.name+'_현금출납장.xlsx'
    created_file = UploadFile.objects.create(title=str(ymd)+'_'+business.name+'_현금출납장', file=upfile, user=request.user)

    wb = load_workbook(filename='./media/'+upfile.name)
    sheet = wb.worksheets[1]

    # 년도 유효성검증
    year = str(sheet['B2'].value).strip()
    if len(year) != 4 or not year.isdigit():
        return HttpResponse("<script>alert('년도 입력형식이 잘못되었습니다. 확인 후 다시 등록해주세요.');history.back();</script>")

    num = TBLBANK.objects.all().order_by('-Bkid').first().Bkid + 1
    error = ""
    for index in range(4, sheet.max_row + 1):
        sindex = str(index)
        isValid = True

        mmdd = str(sheet['A' + sindex].value)
        jukyo = sheet['B' + sindex].value
        item_context = sheet['C' + sindex].value
        input = sheet['D' + sindex].value
        output = sheet['E' + sindex].value
        jango = sheet['F' + sindex].value

        # 필수입력 체크
        if mmdd == None or jukyo == None or item_context == None \
                or input == None or output == None or jango == None:
            error += "[" + sindex + "행] 입력항목 누락\\n"
            continue

        # 날짜 유효성검증
        try:
            datetimecell = year + "/" + mmdd
            Bkdate = datetime.datetime.strptime(datetimecell, '%Y/%m/%d')
        except Exception as e:
            print(e)
            isValid = False
            error += "[" + sindex + "행] 날짜형식 오류\\n"

        # 적요 유효성검증
        # 100자 이상 넘어갈 경우 100자까지 잘라서 저장하도록 변경
        if len(jukyo) > 100:
            # isValid = False
            # error += "[" + sindex + "행] 적요 100자 초과\\n"
            jukyo = jukyo[:100]

        # 관항목 유효성검증 (목명칭으로 검증)
        try:
            month = str(Bkdate.month)
            sessionInfo = session_info(year, month, business.session_month)
            sessionYear = sessionInfo['year']
            item = Item.objects.get(
                paragraph__subsection__institution=business.type3,
                paragraph__subsection__year=sessionYear,
                context=item_context.replace(" ", "").replace(".", "·"))
        except Exception as e:
            print(e)
            isValid = False
            error += "[" + sindex + "행] 관항목 매칭오류\\n"

        # 금액 유효성검증 (수입/지출금액은 둘 중 하나만 있어야 함)
        if input and not output:
            Bkinput = input
            Bkoutput = 0
        elif output and not input:
            Bkinput = 0
            Bkoutput = output
        else:
            isValid = False
            error += "[" + sindex + "행] 금액입력 오류\\n"

        if isValid:
            try:
                with transaction.atomic():
                    TBLBANK.objects.create(
                        Bkid = num,
                        Bkdivision = 1,
                        Mid = request.user.username,
                        Bkacctno = None,
                        Bkname = None,
                        Bkdate = Bkdate,
                        Bkjukyo = jukyo,
                        Bkinput = Bkinput,
                        Bkoutput = Bkoutput,
                        Bkjango = jango,
                        business = business,
                        direct = True
                    )

                    Transaction.objects.create(
                        Bkid = num,
                        Bkdivision = 1,
                        Mid = request.user.username,
                        Bkacctno = None,
                        Bkname = None,
                        Bkdate = Bkdate,
                        Bkjukyo = jukyo,
                        Bkinput = Bkinput,
                        Bkoutput = Bkoutput,
                        Bkjango = jango,
                        business = business,
                        regdatetime = today,
                        item = item
                    )
                num += 1
            except Exception as e:
                print (e)
                error += "[" + sindex + "행] 거래내역 등록오류\\n"
    print(error)
    if error != "":
        error += "위 오류항목들은 확인하여 다시 등록해주시기 바랍니다."
        return HttpResponse("<script>alert('"+error+"');window.close(); window.opener.parent.location.reload(); window.parent.location.href='/';</script>")
    else:
        created_file.delete()   # 오류항목 없는 경우 등록한 파일 삭제
        return HttpResponse('<script type="text/javascript">window.close(); window.opener.parent.location.reload(); window.parent.location.href="/";</script>')

@login_required(login_url='/')
def tr_syn(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    month = request.GET.get('month')
    acctid = request.GET.get('acctid')
    year = request.GET.get('year')
    month = request.GET.get('month')
    page = request.GET.get('page')
    page2 = request.GET.get('page2')
    update_list = Transaction.objects.filter(business=business, Bkdate__year = year, Bkdate__month = month).order_by('Bkdate', 'id')

    first_tr = update_list.first()
    jango = first_tr.Bkjango

    for idx, update in enumerate(update_list):
        print(update.Bkdate, update.Bkinput, update.Bkoutput, update.Bkjango)
        if idx != 0:
            if update.Bkinput != 0 :
                    jango = jango + update.Bkinput
            elif update.Bkoutput != 0 :
                    jango = jango - update.Bkoutput
            print(jango)
            update.Bkjango = jango
            update.save()

    response = redirect('transaction_history')
    response['Location'] += '?page='+page+'&page2='+page2+'&year='+year+'&month='+month+'&acctid='+acctid
    return response


@login_required(login_url='/')
def test(request):
    return render(request,'accounting/test.html')

@login_required(login_url='/')
def design_test(request):
    return render(request,'accounting/design_test.html')

def check_date(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    year2 = int(request.POST.get('year2'))
    month2 = int(request.POST.get('month2'))
    error_message = None

    if business.type3_id == "어린이집" :
        if year > year2:
            year2 = year
            month2 = month
        elif year == year2:
            if month < 3:
                if month2 >= 3:
                    month2 = 2
                elif month > month2:
                    month2 = month
            else:
                if month > month2:
                    month2 = month
                if month2 < 3:
                    month2 = 12
        else:
            if year2 - year > 1:
                year2 = year+1
            if month < 3:
                year2 = year
                month2 = month
            else:
                if month2 > 2:
                    month2 = 2
    else:
        if year > year2:
            year2 = year
            if month > month2:
                month2 = month
        elif year == year2:
            if month > month2:
                month2 = month
        else:
            year2 = year
            if month > month2:
                month2 = month

    context = {'year': year, 'month': month, 'year2': year2, 'month2': month2, 'error_message': error_message}
    return HttpResponse(json.dumps(context), content_type="application/json")


def readNumber(num):
    isMinus = True if num < 0 else False;
    strNum = str(abs(num))

    # 만 단위 자릿수
    tenThousandPos = 4
    # 억 단위 자릿수
    hundredMillionPos = 9
    txtDigit = ['', '십', '백', '천', '만', '억']
    txtNumber = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
    txtPoint = '쩜 '
    resultStr = ''
    #자릿수 카운트
    digitCount = len(strNum) - 1

    index = 0
    while index < len(strNum):
        showDigit = True
        ch = strNum[index]

        #----------숫자표기----------
        #자릿수가 2자리이상이고 1이면 '일'은 표시 안함.
        # 단 '만' '억'에서는 표시 함
        if(digitCount > 1) and (digitCount != tenThousandPos) and  (digitCount != hundredMillionPos) and int(ch) == 1:
            if index == 0 :
                resultStr = resultStr + txtNumber[int(ch)]
            else:
                resultStr = resultStr + ''
        #0이면 숫자표기 안함
        elif int(ch) == 0:
            resultStr = resultStr + ''
            #단 '만,'억'에서는 단위표시해줌
            if (digitCount != tenThousandPos) and  (digitCount != hundredMillionPos):
                showDigit = False
        else:
            resultStr = resultStr + txtNumber[int(ch)]

        #----------단위표기----------
        # 1억 이상
        if digitCount > hundredMillionPos:
            if showDigit:
                resultStr = resultStr + txtDigit[digitCount-hundredMillionPos]
        # 1만 이상
        elif digitCount > tenThousandPos:
            if showDigit:
                resultStr = resultStr + txtDigit[digitCount-tenThousandPos]
        else:
            if showDigit:
                resultStr = resultStr + txtDigit[digitCount]

        digitCount = digitCount - 1
        index = index + 1

    # ----------마이너스표기----------
    resultStr = "-" + resultStr if isMinus else resultStr

    return resultStr


@login_required
def set_proofnum(request):
    ym_list = Transaction.objects.exclude(Bkdivision=0).values('business','item__paragraph__subsection__type','Bkdate__year','Bkdate__month').distinct()
    #print(ym_list)
    for ym in ym_list:
        index = 1
        tr_list = Transaction.objects.filter(business=ym['business'], item__paragraph__subsection__type=ym['item__paragraph__subsection__type'], Bkdate__year=ym['Bkdate__year'], Bkdate__month=ym['Bkdate__month']).exclude(Bkdivision=0).order_by('id')
        for tr in tr_list:
            #print(ym['business'], index, tr)
            tr.proofnum = index
            tr.save()
            index += 1
        #print(index-1)

    return render(request, 'accounting/other_settings.html')


'''
@login_required
def premonth_transfer_price(request):
    Bkid = request.POST.get('Bkid')
    Bkinfo = Account.objects.get(id=Bkid)
    business = Business.objects.get(id=1)
    today = datetime.datetime.now()
    this_month = DateFormat(today).format("Y-m")+'-01'
    one_month_ago = DateFormat(today - relativedelta(months=1)).format("Y-m")+'-01'


    if request.method == "POST":
        data = TBLBANK.objects.filter(Bkacctno=Bkinfo.account_number.replace('-','')).filter(Bkdate__gte=one_month_ago).filter(Bkdate__lt=this_month).order_by('-Bkdate','-Bkid').first()
    
    is_regist = True

    try :
        transaction = Transaction.objects.filter(Bkacctno=Bkinfo.account_number).filter(Bkdate__gte=this_month).get(Bkdivision=0)
    except IndexError:
        return HttpResponse("<script>alert('해당계좌의 거래내역이 없습니다.');history.back();</script>")
    except Transaction.DoesNotExist:
        transaction = Transaction(
            Bkid=data.Bkid,
            Bkdivision=0,
            Mid=request.user.username,
            Bkacctno=Bkinfo.account_number,
            Bkname=Bkinfo.bank.name,
            Bkdate=this_month,
            Bkjukyo="전월이월금",
            Bkinput=data.Bkjango,
            Bkoutput=0,
            Bkjango=data.Bkjango,
            regdatetime=today,
        )
        transaction.save()
        is_regist = False

    if is_regist:
        return HttpResponse("<script>alert('해당계좌의 전월이월금은 입력된 상태입니다. 계좌를 확인하고 다시 등록해주세요.');history.back();</script>")
    
    return redirect('transaction_history', business.pk)
    #return HttpResponse(json.dumps(response), content_type='application/json')
'''

@login_required
def budget_spi_total(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    year = request.POST.get('year')
    type = request.POST.get('type')
    context = request.POST.get('context')

    total = Budget.objects.filter(
        business = business, year = year, type=type
    ).filter(
        Q(item__context__contains=context)
        | Q(item__paragraph__context__contains=context)
        | Q(item__paragraph__subsection__context__contains=context)
    ).aggregate(
        total=Coalesce(Sum('price'), 0)
    )['total']

    context = {'total': total}
    return HttpResponse(json.dumps(context), content_type="application/json")

@login_required(login_url='/')
def erowm_account(request):
    accounts = Account.objects.all()

    for list in accounts:
        data = {'bkacctno': list.account_number}
        list.bankda = account_info_xml(data)

    return render(request, 'accounting/erowm_account.html', {
        'bankda_page': 'active', 'erowm_account_page': 'active', 'accounts': accounts
    })


# ajax
from django.http.response import JsonResponse

def subsection_list(request):
    if request.method == "POST":
        institution = request.POST.get('institution')
        year = request.POST.get('year')
        type = request.POST.get('type')

        subsection = getSubsectionList(institution, year, type)   # QuerySet
        data = list(subsection.values())     # JsonResponse를 사용하여 전달하기 위해 QuerySet을 list 타입으로 변경

        return JsonResponse(data, safe=False)   # safe=False 필수

def paragraph_list(request):
    if request.method == "POST":
        subsection = request.POST.get('subsection')

        paragraph = getParagraphList(subsection)   # QuerySet
        data = list(paragraph.values())     # JsonResponse를 사용하여 전달하기 위해 QuerySet을 list 타입으로 변경

        return JsonResponse(data, safe=False)   # safe=False 필수

def item_list(request):
    if request.method == "POST":
        paragraph = request.POST.get('paragraph')

        item = getItemList(paragraph)  # QuerySet
        data = list(item.values())  # JsonResponse를 사용하여 전달하기 위해 QuerySet을 list 타입으로 변경

        return JsonResponse(data, safe=False)  # safe=False 필수

@login_required(login_url='/')
def regist_transaction(request):
    business = get_object_or_404(Business, pk=request.session['business'])
    acctid = request.POST.get('acctid')
    acct = get_object_or_404(Account, business=business, id=acctid)
    Mid = request.user.username
    tr_list = request.POST.get('transaction_list').split(",")
    Bkdate_list = request.POST.get('Bkdate_list').split(",")
    Bkjukyo_list = request.POST.get('Bkjukyo_list').split(",")
    item_list = request.POST.get('item_list').split(",")
    subdivision_list = request.POST.get('subdivision_list').split(",")
    relative_subsection_list = request.POST.get('input_subsection_list').split(",")
    relative_item_list = request.POST.get('input_subdivision_list').split(",")

    if request.method == "POST":
        # from django.db import transaction
        # with transaction.atomic():
        try:
            for index, tr in enumerate(tr_list):
                Bkid = tr_list[index]
                tblbank_tr = TBLBANK.objects.get(Bkid=Bkid)
                try:
                    subdivision = Subdivision.objects.get(id=subdivision_list[index])
                except Exception as e:
                    print(e)
                    subdivision = None
                try:
                    relative_subsection = Subsection.objects.get(id=relative_subsection_list[index])
                except Exception as e:
                    print(e)
                    relative_subsection = None
                try:
                    relative_item = Item.objects.get(id=relative_item_list[index])
                except Exception as e:
                    print(e)
                    relative_item = None

                tr = Transaction(
                    Bkid=Bkid,
                    Mid=Mid,
                    business=business,
                    Bkacctno=acct.account_number,
                    Bkname=acct.bank.name,
                    Bkdate=datetime.datetime.strptime(Bkdate_list[index], "%Y-%m-%d"),
                    Bkjukyo=Bkjukyo_list[index],
                    Bkinput=tblbank_tr.Bkinput,
                    Bkoutput=tblbank_tr.Bkoutput,
                    item=Item.objects.get(id=item_list[index]),
                    subdivision=subdivision,
                    relative_subsection=relative_subsection,
                    relative_item=relative_item
                )
                registTransaction(business, tr)
        except DeadlineCompletionError as e:
            print(e.__str__())
            return JsonResponse(e.__str__(), safe=False)  # safe=False 필수
        except NoTransactionHistoryForPreviousMonth as e:
            print(e.__str__())
            return JsonResponse(e.__str__(), safe=False)  # safe=False 필수
    return JsonResponse("SUCCESS", safe=False)  # safe=False 필수